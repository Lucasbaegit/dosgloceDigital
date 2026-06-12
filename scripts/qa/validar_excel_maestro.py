"""Generate and validate the polished master price workbook."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
import json
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "src"))

from export import PricesExcelExporter, PricesTablesBuilder  # noqa: E402


REQUIRED_SHEETS = PricesExcelExporter.REQUIRED_SHEETS
VARIABLE_HEADERS = [
    "key",
    "label",
    "value",
    "unit",
    "editable",
    "editable_en_sistema",
    "editable_en_excel_maestro",
    "tipo",
    "impacta_hoy",
    "estado_operativo",
    "productos_afectados",
    "source_file",
    "source_path",
    "source_sheet",
    "source_cell",
    "description",
]
BROKEN_PATTERNS = ["�", "m?nimos", "m?nimo", "ilustraci?n", "f?rmula", "f?rmulas", "configuraci?n", "informaci?n", "hist?rico", "l?gica", "Im?n"]
FORBIDDEN_VARIABLE_KEYS = ("precio_final", "precio_pdf", "precio_objetivo_pdf", "factor_ajuste_pdf", "total_final", "matriz_pdf")
OPERATIVE_KEYS = ["tipo_cambio_usd", "click_color", "obra_90g", "multiplicador_general", "adicional_tinta_blanca_base_1_copia"]
PRICE_TRACE_HEADERS = [
    "producto",
    "familia",
    "variante",
    "material",
    "gramaje",
    "formato",
    "cantidad",
    "rango",
    "caras",
    "terminacion",
    "adicional",
    "modo_precio",
    "precio_final",
    "precio_unitario",
    "componente",
    "tipo_componente",
    "variable_key",
    "variable_label",
    "valor_base",
    "unidad",
    "operacion",
    "factor_multiplicador",
    "subtotal_componente",
    "fuente_componente",
    "fuente_precio_final",
    "editable_en_sistema",
    "editable_en_excel_maestro",
    "impacta_hoy",
    "estado_operativo",
    "observacion",
]


def main() -> int:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_dir = ROOT / "reports" / "exports"
    report_dir.mkdir(parents=True, exist_ok=True)
    report_path = report_dir / f"validacion_excel_maestro_{timestamp}.json"

    results: list[dict[str, Any]] = []
    export = PricesTablesBuilder(ROOT).build()
    filename, payload = PricesExcelExporter(
        ROOT,
        validation_excel_status="ok",
        validation_report_path=str(report_path.resolve()),
    ).render(export)
    workbook = load_workbook(BytesIO(payload), data_only=True)

    check(results, "workbook_openpyxl", True, "Workbook abre correctamente con openpyxl.")
    check(results, "filename", filename.startswith("cotizador_digital_maestro_") and filename.endswith(".xlsx"), filename)
    check(results, "required_sheets", workbook.sheetnames == REQUIRED_SHEETS, workbook.sheetnames)
    check(results, "summary_fields", summary_has_required_fields(workbook["00_RESUMEN"]), "Resumen contiene version, fuentes, dirty y conteos.")
    summary = summary_values(workbook["00_RESUMEN"])
    operative_count = count_operative_variables(workbook["01_VARIABLES_MADRE"])
    check(results, "summary_validacion_excel_ok", summary.get("validacion_excel") == "ok", summary.get("validacion_excel"))
    check(results, "summary_variables_operativas", summary.get("variables_madre_operativas") == operative_count, {"summary": summary.get("variables_madre_operativas"), "real": operative_count})
    broken = find_broken_text(workbook)
    check(results, "caracteres_rotos", not broken, broken)
    check(results, "variable_headers", [cell.value for cell in workbook["01_VARIABLES_MADRE"][1]] == VARIABLE_HEADERS, [cell.value for cell in workbook["01_VARIABLES_MADRE"][1]])
    check(results, "variables_no_final_prices", variables_do_not_include_final_prices(workbook["01_VARIABLES_MADRE"]), "Variables madre sin precios finales ni factor_ajuste_pdf.")
    operative_validation = validate_operative_variables(workbook["01_VARIABLES_MADRE"])
    check(results, "variables_operativas", operative_validation["ok"], operative_validation)
    prepared_validation = validate_prepared_variables(workbook["01_VARIABLES_MADRE"])
    check(results, "variables_preparadas", prepared_validation["ok"], prepared_validation)
    check(results, "ranges_fixed", ranges_are_fixed(workbook["02_RANGOS"]), "Rangos no editables.")
    blocked_validation = validate_blocked(workbook["18_BLOQUEADOS"])
    check(results, "bloqueados", blocked_validation["ok"], blocked_validation)
    trace_validation = validate_trace(workbook["19_TRAZABILIDAD"])
    check(results, "trazabilidad", trace_validation["ok"], trace_validation)
    price_trace_validation = validate_price_trace(workbook["21_TRAZABILIDAD_PRECIOS"])
    check(results, "trazabilidad_precios", price_trace_validation["ok"], price_trace_validation)
    formatting_validation = validate_basic_format(workbook)
    check(results, "formato_basico", formatting_validation["ok"], formatting_validation)

    price_results = validate_prices(workbook)
    for item in price_results:
        check(results, item["check"], item["ok"], item)

    failed = [item for item in results if item["status"] != "PASS"]
    report = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "archivo_excel": str((report_dir / filename).resolve()),
        "archivo_bytes": len(payload),
        "checks_totales": len(results),
        "checks_pass": len(results) - len(failed),
        "checks_fail": len(failed),
        "hojas_encontradas": workbook.sheetnames,
        "precios_testigo_validados": [item for item in results if item["check"].startswith("precio_") and item["status"] == "PASS"],
        "caracteres_rotos_detectados": broken,
        "variables_operativas_validadas": operative_validation,
        "variables_preparadas_validadas": prepared_validation,
        "bloqueados_validados": blocked_validation,
        "trazabilidad_precios_validada": price_trace_validation,
        "veredicto": "PASS" if not failed else "FAIL",
        "resultados": results,
    }
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({"status": report["veredicto"], "excel": report["archivo_excel"], "reporte": str(report_path.resolve()), "fallos": len(failed)}, ensure_ascii=False))
    return 1 if failed else 0


def check(results: list[dict[str, Any]], name: str, passed: bool, detail: Any) -> None:
    results.append({"check": name, "status": "PASS" if passed else "FAIL", "detail": detail})


def sheet_contains(ws, expected: str) -> bool:
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if str(value) == expected:
                return True
    return False


def summary_has_required_fields(ws) -> bool:
    required = {
        "fecha_generacion",
        "version_base_git",
        "commit_actual",
        "repo_dirty",
        "nota_version",
        "fuente_precios_finales",
        "fuente_logica_historica",
        "precios_sin_iva",
        "tablas_pdf_fijas_no_editables",
        "variables_madre_operativas",
        "variables_madre_preparadas",
        "productos_activos",
        "productos_bloqueados",
        "hojas_generadas",
        "validacion_excel",
        "trazabilidad_precios",
        "hoja_trazabilidad_precios",
    }
    found = {str(row[0]) for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    return required <= found


def summary_values(ws) -> dict[str, Any]:
    return {str(row[0]): row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}


def count_operative_variables(ws) -> int:
    rows = variable_rows(ws)
    return sum(1 for row in rows.values() if row.get("estado_operativo") == "operativa")


def find_broken_text(workbook) -> list[dict[str, Any]]:
    found = []
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                value = cell.value
                if any(pattern in value for pattern in BROKEN_PATTERNS):
                    found.append({"sheet": ws.title, "cell": cell.coordinate, "value": value})
    return found


def variable_rows(ws) -> dict[str, dict[str, Any]]:
    headers = [cell.value for cell in ws[1]]
    return {row[0]: dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}


def variables_do_not_include_final_prices(ws) -> bool:
    for key in variable_rows(ws):
        if key == "rango" or any(token in str(key) for token in FORBIDDEN_VARIABLE_KEYS):
            return False
    return True


def validate_operative_variables(ws) -> dict[str, Any]:
    rows = variable_rows(ws)
    details = {}
    ok = True
    for key in OPERATIVE_KEYS:
        row = rows.get(key)
        valid = bool(row) and row["editable_en_sistema"] is True and row["impacta_hoy"] is True and row["estado_operativo"] == "operativa"
        details[key] = row
        ok = ok and valid
    return {"ok": ok, "detalles": details}


def validate_prepared_variables(ws) -> dict[str, Any]:
    rows = variable_rows(ws)
    prepared = [row for row in rows.values() if row["impacta_hoy"] is False]
    ok = bool(prepared) and all(row["editable_en_sistema"] is False and row["editable_en_excel_maestro"] is True and row["estado_operativo"] == "preparada_no_conectada" for row in prepared)
    return {"ok": ok, "cantidad": len(prepared), "ejemplos": prepared[:3]}


def ranges_are_fixed(ws) -> bool:
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    editable_index = headers.index("editable")
    tipo_index = headers.index("tipo")
    return all(row[editable_index] is False and row[tipo_index] == "rango_fijo" for row in rows[1:] if row[0])


def validate_blocked(ws) -> dict[str, Any]:
    headers = [cell.value for cell in ws[1]]
    rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    required = ["Membretes", "Terminaciones Tarjetas", "OPP Stickers Circulares", "DTF UV", "DTF Textil", "PegaManía"]
    present = {row["producto"] for row in rows}
    ok = all(name in present for name in required) and all(row.get("precio_inventado") is False for row in rows)
    return {"ok": ok, "presentes": sorted(present), "cantidad": len(rows)}


def validate_trace(ws) -> dict[str, Any]:
    headers = [cell.value for cell in ws[1]]
    required_headers = ["producto", "endpoint", "modo_precio", "fuente_pdf", "fuente_excel", "archivo_data", "motor_backend", "estado", "notas"]
    rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    modes = {row.get("modo_precio") for row in rows}
    ok = headers == required_headers and "/bajadas-v2/cotizar" in {row.get("endpoint") for row in rows} and "formula_editable_calibrada" in modes and "tabla_fija_pdf" in modes and "bloqueado" in modes
    return {"ok": ok, "headers": headers, "modos": sorted(str(mode) for mode in modes), "cantidad": len(rows)}


def validate_price_trace(ws) -> dict[str, Any]:
    headers = [cell.value for cell in ws[1]]
    rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    required_cases = {
        "tarjetas_9x5_5139": any(row.get("producto") == "Tarjetas 9x5" and row.get("precio_final") == 5139 for row in rows),
        "stickers_circulares_85980": any(row.get("producto") == "Stickers Circulares" and row.get("precio_final") == 85980 for row in rows),
        "autoadhesivas_tinta_603": any(row.get("producto") == "Bajadas Autoadhesivas" and row.get("componente") == "tinta_blanca" and row.get("valor_base") == 603 for row in rows),
        "bajadas_fullcolor_622": any(row.get("producto") == "Bajadas Fullcolor" and row.get("precio_unitario") == 622 for row in rows),
        "membretes_bloqueado": any(row.get("producto") == "Membretes" and row.get("modo_precio") == "bloqueado" for row in rows),
    }
    tipos = {row.get("tipo_componente") for row in rows}
    componentes = {row.get("componente") for row in rows}
    tabla_pdf_safe = all(
        row.get("editable_en_sistema") is not True and row.get("impacta_hoy") is not True
        for row in rows
        if row.get("tipo_componente") == "tabla_pdf" or row.get("estado_operativo") == "tabla_pdf_fija"
    )
    blocked_safe = all(
        row.get("precio_final") in (None, "")
        and row.get("precio_unitario") in (None, "")
        and row.get("subtotal_componente") in (None, "")
        for row in rows
        if row.get("modo_precio") == "bloqueado" or row.get("tipo_componente") == "bloqueado"
    )
    ok = (
        headers == PRICE_TRACE_HEADERS
        and all(required_cases.values())
        and "tabla_pdf" in tipos
        and "variable_madre" in tipos
        and "adicional_variable" in tipos
        and ({"factor", "multiplicador"} & tipos)
        and "total_final" in componentes
        and tabla_pdf_safe
        and blocked_safe
    )
    return {
        "ok": ok,
        "headers": headers,
        "cantidad": len(rows),
        "required_cases": required_cases,
        "tipos": sorted(str(tipo) for tipo in tipos),
        "componentes": sorted(str(componente) for componente in componentes),
        "tabla_pdf_safe": tabla_pdf_safe,
        "blocked_safe": blocked_safe,
    }


def validate_basic_format(workbook) -> dict[str, Any]:
    failures = []
    for ws in workbook.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            failures.append({"sheet": ws.title, "motivo": "hoja_vacia"})
        if ws.freeze_panes != "A2":
            failures.append({"sheet": ws.title, "motivo": "freeze_panes"})
        if not ws.auto_filter.ref:
            failures.append({"sheet": ws.title, "motivo": "autofilter"})
        if not any(cell.font.bold for cell in ws[1]):
            failures.append({"sheet": ws.title, "motivo": "headers_no_bold"})
        if not any(dim.width and dim.width >= 12 for dim in ws.column_dimensions.values()):
            failures.append({"sheet": ws.title, "motivo": "anchos_columnas"})
    return {"ok": not failures, "fallos": failures}


def validate_prices(workbook) -> list[dict[str, Any]]:
    expected_rows = [
        ("03_BAJADAS_FULLCOLOR", {"categoria": "Bajadas Kraft", "cantidad_rango": "1", "gramaje": "80g", "caras": "4/0"}, "precio_sin_iva", 782),
        ("03_BAJADAS_FULLCOLOR", {"categoria": "Bajadas Kraft", "cantidad_rango": "51 a 100", "gramaje": "80g", "caras": "4/0"}, "precio_sin_iva", 569),
        ("03_BAJADAS_FULLCOLOR", {"categoria": "Bajadas Fullcolor", "cantidad_rango": "51 a 100", "gramaje": "150g", "caras": "4/0"}, "precio_sin_iva", 622),
        ("05_AUTOADHESIVAS", {"adicional": "laca_uv", "rango": "51 a 100"}, "precio_unitario_sin_iva", 106),
        ("05_AUTOADHESIVAS", {"adicional": "laca_uv", "rango": "26 a 50"}, "precio_unitario_sin_iva", 116),
        ("05_AUTOADHESIVAS", {"adicional": "tinta_blanca"}, "precio_unitario_sin_iva", 603),
        ("09_TARJETAS_9X5", {"cantidad_unidades": 100, "terminacion": "sin_laminar", "caras": "4/0"}, "precio_total_sin_iva", 5139),
        ("09_TARJETAS_9X5", {"cantidad_unidades": 1000, "terminacion": "laminado_mate", "caras": "4/4"}, "precio_total_sin_iva", 48401),
        ("10_TARJETAS_POSTALES", {"cantidad_unidades": 100, "terminacion": "sin_laminar", "caras": "4/0"}, "precio_total_sin_iva", 10932),
        ("10_TARJETAS_POSTALES", {"cantidad_unidades": 1000, "terminacion": "laminado_mate", "caras": "4/4"}, "precio_total_sin_iva", 136795),
        ("11_FOLLETOS", {"cantidad_unidades": 1000, "formato": "A4", "gramaje": "80g", "caras": "1/1"}, "precio_total_sin_iva", 119247),
        ("11_FOLLETOS", {"cantidad_unidades": 100, "formato": "10x15", "gramaje": "150g", "caras": "4/0"}, "precio_total_sin_iva", 8445),
        ("11_FOLLETOS", {"cantidad_unidades": 100, "formato": "10x10", "gramaje": "150g", "caras": "4/0"}, "precio_total_sin_iva", 6496),
        ("12_CARPETAS", {"cantidad_rango": "1", "terminacion": "sin_laminar", "caras": "4/0"}, "precio_unitario", 1762),
        ("12_CARPETAS", {"cantidad_rango": "1", "terminacion": "sin_laminar", "caras": "4/0"}, "precio_solapa_impresa_unitario", 255),
        ("12_CARPETAS", {"cantidad_rango": "51 a 100", "terminacion": "laca_uv", "caras": "4/4"}, "precio_unitario", 1683),
        ("12_CARPETAS", {"cantidad_rango": "51 a 100", "terminacion": "laca_uv", "caras": "4/4"}, "precio_solapa_impresa_unitario", 188),
        ("13_SOBRES", {"cantidad_unidades": 100, "tipo_sobre": "sobre_bolsa_a4_22_9x32_4"}, "precio_unitario", 633),
        ("13_SOBRES", {"cantidad_unidades": 1000, "tipo_sobre": "sobre_bolsa_27x37"}, "precio_unitario", 536),
        ("14_STICKERS_CORTE_RECTO", {"cantidad_unidades": 100, "formato": "6x4", "terminacion": "sin_laca_uv"}, "precio_total_sin_iva", 2765),
        ("14_STICKERS_CORTE_RECTO", {"cantidad_unidades": 1000, "formato": "10x7", "terminacion": "con_laca_uv"}, "precio_total_sin_iva", 61703),
        ("15_IMANES_CORTE_RECTO", {"cantidad_unidades": 100, "formato": "6x4", "terminacion": "sin_laca_uv"}, "precio_total_sin_iva", 7526),
        ("15_IMANES_CORTE_RECTO", {"cantidad_unidades": 1000, "formato": "10x7", "terminacion": "con_laca_uv"}, "precio_total_sin_iva", 153680),
        ("07_STICKERS_CIRCULARES", {"cantidad_unidades": 1000, "formato": "10cm", "terminacion": "con_laca_uv"}, "precio_total_sin_iva", 85980),
        ("16_PLANCHA_IMAN", {"cantidad_rango": "301 a 500"}, "precio_unitario_sin_iva", 1931),
        ("17_AGENDAS_CUADERNOS", {"producto": "agenda_2026", "formato": "A5", "paginas": 72}, "precio_unitario_sin_iva", 3000),
    ]
    results = []
    for sheet_name, criteria, value_column, expected in expected_rows:
        actual = find_table_value(workbook[sheet_name], criteria, value_column)
        results.append({
            "check": f"precio_{sheet_name}_{value_column}_{expected}",
            "ok": actual == expected,
            "sheet": sheet_name,
            "criteria": criteria,
            "value_column": value_column,
            "expected": expected,
            "actual": actual,
        })
    return results


def find_table_value(ws, criteria: dict[str, Any], value_column: str) -> Any:
    rows = list(ws.iter_rows(values_only=True))
    for index, row in enumerate(rows):
        headers = [str(value) if value is not None else "" for value in row]
        if value_column not in headers:
            continue
        if not all(key in headers for key in criteria):
            continue
        indexes = {header: headers.index(header) for header in headers if header}
        for candidate in rows[index + 1:]:
            if not any(value is not None for value in candidate):
                break
            if all(candidate[indexes[key]] == expected for key, expected in criteria.items()):
                return candidate[indexes[value_column]]
    return None


if __name__ == "__main__":
    raise SystemExit(main())
