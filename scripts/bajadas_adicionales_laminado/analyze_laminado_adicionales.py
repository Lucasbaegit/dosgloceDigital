import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / 'original' / 'Copia de DIGITAL sistema 2025.xlsx'
DATA_DIR = ROOT / 'data' / 'bajadas_adicionales_laminado'
DOCS_DIR = ROOT / 'docs' / 'bajadas_adicionales_laminado'


def find_used_range(ws):
    min_r=min_c=10**9
    max_r=max_c=0
    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            v=ws.cell(r,c).value
            if v not in (None, ''):
                min_r=min(min_r,r); min_c=min(min_c,c); max_r=max(max_r,r); max_c=max(max_c,c)
    if max_r == 0:
        return None
    return {
        'min_row': min_r,
        'max_row': max_r,
        'min_col': min_c,
        'max_col': max_c,
        'range': f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}",
    }


def c(ws, addr):
    return ws[addr].value


def data_value(wsd, addr):
    return wsd[addr].value


def parse_scale_label(label):
    s = str(label).strip() if label is not None else ''
    if s == '1':
        return {'raw': s, 'from': 1, 'to': 1}
    if '+' in s:
        try:
            n = int(''.join(ch for ch in s if ch.isdigit()))
            return {'raw': s, 'from': n, 'to': None}
        except Exception:
            return {'raw': s, 'from': None, 'to': None}
    if '-' in s:
        a, b = [p.strip() for p in s.split('-', 1)]
        if a.isdigit() and b.isdigit():
            return {'raw': s, 'from': int(a), 'to': int(b)}
    return {'raw': s, 'from': None, 'to': None}


def main():
    wb = load_workbook(XLSX, data_only=False, read_only=False)
    wbd = load_workbook(XLSX, data_only=True, read_only=False)
    ws = wb['Laminado']
    wsd = wbd['Laminado']

    used = find_used_range(ws)

    scales = []
    for r in range(7, 14):
        scales.append({
            'row': r,
            'label_cell': f'B{r}',
            'label': c(ws, f'B{r}'),
            'parsed': parse_scale_label(c(ws, f'B{r}')),
        })

    # Additional blocks
    laca_rows = []
    brillo_rows = []
    mate_rows = []
    for r in range(7, 14):
        laca_rows.append({
            'row': r,
            'formula_cell': f'E{r}',
            'formula': c(ws, f'E{r}'),
            'value': data_value(wsd, f'E{r}'),
        })
        brillo_rows.append({
            'row': r,
            'formula_cell': f'C{r}',
            'formula': c(ws, f'C{r}'),
            'value': data_value(wsd, f'C{r}'),
        })
        mate_rows.append({
            'row': r,
            'formula_cell': f'D{r}',
            'formula': c(ws, f'D{r}'),
            'value': data_value(wsd, f'D{r}'),
        })

    formulas = []
    for r in range(used['min_row'], used['max_row']+1):
        for col in range(used['min_col'], used['max_col']+1):
            addr = f"{get_column_letter(col)}{r}"
            v = c(ws, addr)
            if isinstance(v, str) and v.startswith('='):
                formulas.append({'cell': addr, 'formula': v, 'value': data_value(wsd, addr)})

    formula_groups = {
        'laminado_brillo': {'pattern': '=C5*Lx*D2', 'cells': [f'C{r}' for r in range(7,14)]},
        'laminado_mate': {'pattern': '=Cx*D2', 'cells': [f'D{r}' for r in range(7,14)]},
        'laca_uv': {'pattern': '=Cx/I2', 'cells': [f'E{r}' for r in range(7,14)]},
        'copias_margen_1_9': {'pattern': '=base*1.9', 'cells': [f'N{r}' for r in range(7,14)] + [f'O{r}' for r in range(7,14)] + [f'P{r}' for r in range(7,14)]},
    }

    dependencies = {
        'internal_cells': [
            'A2', 'A3', 'A4', 'B1', 'B4', 'C1', 'C2', 'D2', 'I2',
            'C5', 'L7:L13', 'C7:C13', 'D7:D13', 'E7:E13'
        ],
        'named_context': [
            {'cell': 'A3', 'meaning': 'Costo base hoja A3+ (32x47cm)'},
            {'cell': 'C2', 'meaning': 'Laminado por lado (factor base)'},
            {'cell': 'D2', 'meaning': 'Coeficiente/margen intermedio'},
            {'cell': 'I2', 'meaning': 'Coeficiente para derivar laca UV'},
            {'cell': 'L7:L13', 'meaning': 'Escala por cantidad para costo unitario decreciente'},
            {'cell': 'C23', 'meaning': 'Dólar de referencia presente en hoja Laminado (no referencia directa a PAPEL US$)'}
        ],
        'external_sheet_references': [],
        'external_file_references': [],
    }

    trace_tree = {
        'laca': {
            'base': 'A3+ por pliego',
            'quantity_scale': 'B7:B13',
            'path': [
                {'node': 'Costo unitario base', 'cell': 'C5', 'formula': '=A2/B4*C4*C1', 'value': data_value(wsd, 'C5')},
                {'node': 'Escala cantidad', 'cell': 'Lx', 'formula': 'valor fijo por tramo', 'values': {f'L{r}': c(ws, f'L{r}') for r in range(7,14)}},
                {'node': 'Aplicación laca', 'cell': 'Ex', 'formula': '=Cx/I2', 'values': {f'E{r}': data_value(wsd, f'E{r}') for r in range(7,14)}},
                {'node': 'Resultado unitario', 'cell': 'E7:E13', 'meaning': 'precio unitario por pliego A3+ según escala'}
            ],
            'dependencies': ['C5', 'L7:L13', 'I2', 'C7:C13'],
        },
        'laminado_brillo': {
            'base': 'A3+ por pliego',
            'quantity_scale': 'B7:B13',
            'path': [
                {'node': 'Costo unitario base', 'cell': 'C5', 'formula': '=A2/B4*C4*C1', 'value': data_value(wsd, 'C5')},
                {'node': 'Escala cantidad', 'cell': 'Lx', 'values': {f'L{r}': c(ws, f'L{r}') for r in range(7,14)}},
                {'node': 'Aplicación brillo', 'cell': 'Cx', 'formula': '=C5*Lx*D2', 'values': {f'C{r}': data_value(wsd, f'C{r}') for r in range(7,14)}},
                {'node': 'Resultado unitario', 'cell': 'C7:C13', 'meaning': 'precio unitario brillo por pliego A3+'}
            ],
            'dependencies': ['C5', 'L7:L13', 'D2'],
        },
        'laminado_mate': {
            'base': 'A3+ por pliego',
            'quantity_scale': 'B7:B13',
            'path': [
                {'node': 'Base brillo', 'cell': 'Cx', 'formula': '=C5*Lx*D2', 'values': {f'C{r}': data_value(wsd, f'C{r}') for r in range(7,14)}},
                {'node': 'Derivación mate', 'cell': 'Dx', 'formula': '=Cx*D2', 'values': {f'D{r}': data_value(wsd, f'D{r}') for r in range(7,14)}},
                {'node': 'Resultado unitario', 'cell': 'D7:D13', 'meaning': 'precio unitario mate por pliego A3+'}
            ],
            'dependencies': ['C7:C13', 'D2'],
        },
    }

    inventory = {
        'excel_file': str(XLSX),
        'sheet': 'Laminado',
        'used_range': used['range'],
        'max_row': ws.max_row,
        'max_col': ws.max_column,
        'key_headers': {
            'brillo': 'C6',
            'mate': 'D6',
            'laca_uv': 'E6',
            'scale_labels': 'B7:B13',
            'a3plus_context': 'B3',
        },
        'scale_rows': scales,
        'additional_blocks': {
            'laca': {'range': 'E7:E13', 'header_cell': 'E6', 'header': c(ws, 'E6')},
            'laminado_brillo': {'range': 'C7:C13', 'header_cell': 'C6', 'header': c(ws, 'C6')},
            'laminado_mate': {'range': 'D7:D13', 'header_cell': 'D6', 'header': c(ws, 'D6')},
        },
        'a3_plus_evidence': [
            {'cell': 'B3', 'value': c(ws, 'B3')},
            {'cell': 'G16', 'value': c(ws, 'G16')},
            {'cell': 'G25', 'value': c(ws, 'G25')},
        ],
    }

    formulas_out = {
        'formula_count': len(formulas),
        'formulas': formulas,
        'formula_groups': formula_groups,
        'laca_formulas': laca_rows,
        'laminado_brillo_formulas': brillo_rows,
        'laminado_mate_formulas': mate_rows,
    }

    dependencies_out = dependencies

    trace_out = trace_tree

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    DOCS_DIR.mkdir(parents=True, exist_ok=True)

    (DATA_DIR / 'laminado_inventory.json').write_text(json.dumps(inventory, ensure_ascii=False, indent=2), encoding='utf-8')
    (DATA_DIR / 'laminado_formulas.json').write_text(json.dumps(formulas_out, ensure_ascii=False, indent=2), encoding='utf-8')
    (DATA_DIR / 'laminado_dependencies.json').write_text(json.dumps(dependencies_out, ensure_ascii=False, indent=2), encoding='utf-8')
    (DATA_DIR / 'laminado_trace_tree.json').write_text(json.dumps(trace_out, ensure_ascii=False, indent=2), encoding='utf-8')

    diag_md = f"""# Diagnóstico hoja Laminado

- Hoja analizada: `Laminado`
- Rango usado detectado: `{used['range']}`
- Bloques relevantes:
  - Laca UV: `E7:E13` (encabezado `E6`)
  - Laminado brillo: `C7:C13` (encabezado `C6`)
  - Laminado mate: `D7:D13` (encabezado `D6`)
- Escalas detectadas en `B7:B13`: {', '.join(str(c(ws, f'B{r}')) for r in range(7,14))}
- Evidencia de pliego A3+: `B3`, `G16`, `G25`

## Observaciones clave
- Brillo y mate comparten estructura, pero **no** tienen el mismo valor final (mate deriva de brillo con otro coeficiente).
- Laca UV usa fórmula propia (`=Cx/I2`) y no es simple copia de brillo/mate.
- No hay referencias directas de fórmula a otras hojas (`!`) ni archivos externos (`[]`).
"""

    laca_md = f"""# Árbol Laca UV

- Encabezado: `E6 = {c(ws, 'E6')}`
- Rango de salida: `E7:E13`
- Fórmula patrón: `=Cx/I2`

## Composición
1. Base unitario: `C5 = {c(ws, 'C5')}` -> valor actual `{data_value(wsd, 'C5')}`
2. Escala cantidad: `L7:L13` -> {', '.join(str(c(ws, f'L{r}')) for r in range(7,14))}
3. Base brillo por rango: `Cx = C5*Lx*D2`
4. Laca UV por rango: `Ex = Cx/I2`

## Salidas actuales por escala
"""
    for r in range(7,14):
        laca_md += f"- `{c(ws, f'B{r}')}` -> `E{r}` = {data_value(wsd, f'E{r}')}\n"

    brillo_md = f"""# Árbol Laminado Brillo

- Encabezado: `C6 = {c(ws, 'C6')}`
- Rango de salida: `C7:C13`
- Fórmula patrón: `=C5*Lx*D2`

## Composición
1. Base unitario: `C5` = {data_value(wsd, 'C5')}
2. Escala cantidad: `Lx`
3. Coeficiente: `D2` = {c(ws, 'D2')}
4. Resultado brillo: `Cx`

## Salidas actuales por escala
"""
    for r in range(7,14):
        brillo_md += f"- `{c(ws, f'B{r}')}` -> `C{r}` = {data_value(wsd, f'C{r}')}\n"

    mate_md = f"""# Árbol Laminado Mate

- Encabezado: `D6 = {c(ws, 'D6')}`
- Rango de salida: `D7:D13`
- Fórmula patrón: `=Cx*D2`

## Composición
1. Brillo por rango: `Cx`
2. Coeficiente: `D2` = {c(ws, 'D2')}
3. Resultado mate: `Dx`

## Salidas actuales por escala
"""
    for r in range(7,14):
        mate_md += f"- `{c(ws, f'B{r}')}` -> `D{r}` = {data_value(wsd, f'D{r}')}\n"

    rec_md = """# Recomendación de modelo (futura integración)

## Opciones de adicional
- `sin_adicional`
- `adicional_laca`
- `adicional_laminado_brillo`
- `adicional_laminado_mate`

## Regla de cálculo propuesta
- `total = (precio_base + adicional_unitario * cantidad_unidades) * factor_urgencia`
- El adicional se calcula **antes** de urgencia.
- No se permite combinar múltiples adicionales en una misma cotización.

## Derivación por formato
- Tomar `A3+` como base de cálculo del adicional (coherente con hoja Laminado).
- Derivar a otros formatos con los mismos factores de conversión usados en Bajadas v2 (misma filosofía de escalado por formato).

## Qué guardar en trazabilidad
- adicional elegido
- rango aplicado
- origen técnico en Excel (`C7:C13` brillo / `D7:D13` mate / `E7:E13` laca)
- fórmula patrón
- coeficientes usados (`D2`, `I2`, `Lx`)
- exclusión de combinaciones múltiples

## Notas
- Esta etapa no integra al motor productivo.
- Tinta blanca, precorte, medio corte y otros acabados quedan fuera de este análisis.
"""

    (DOCS_DIR / '01_diagnostico_laminado.md').write_text(diag_md, encoding='utf-8')
    (DOCS_DIR / '02_arbol_laminado_laca.md').write_text(laca_md, encoding='utf-8')
    (DOCS_DIR / '03_arbol_laminado_brillo.md').write_text(brillo_md, encoding='utf-8')
    (DOCS_DIR / '04_arbol_laminado_mate.md').write_text(mate_md, encoding='utf-8')
    (DOCS_DIR / '05_recomendacion_modelo_adicionales.md').write_text(rec_md, encoding='utf-8')

    print('OK: análisis Laminado generado')


if __name__ == '__main__':
    main()
