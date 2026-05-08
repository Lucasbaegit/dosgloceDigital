# -*- coding: utf-8 -*-
import csv
import json
import re
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parents[1]
ORIGINAL = BASE / "original" / "Copia de DIGITAL sistema 2025.xlsx"
DATA = BASE / "data" / "bajadas_autoadhesivas"
DOCS = BASE / "docs" / "bajadas_autoadhesivas"

TARGETS = [
    ("1", 1069, 1797),
    ("2 a 25", 875, 1470),
    ("26 a 50", 826, 1389),
    ("51 a 100", 778, 1307),
    ("101 a 300", 729, 1225),
    ("301 a 500", 680, 1144),
    ("501 a 1000", 632, 1062),
]

CELL_REF_RE = re.compile(r"(?:'([^']+)'!)?\$?([A-Z]{1,3})\$?(\d+)")


def parse_refs(formula):
    refs = []
    if not isinstance(formula, str) or not formula.startswith("="):
        return refs
    for m in CELL_REF_RE.finditer(formula):
        sheet = m.group(1) or "Bajadas"
        cell = f"{m.group(2)}{m.group(3)}"
        refs.append((sheet, cell))
    out, seen = [], set()
    for r in refs:
        if r not in seen:
            out.append(r)
            seen.add(r)
    return out


def calc_diff(excel, target):
    if excel in (None, 0):
        return None, None
    da = excel - target
    dp = (da / target) * 100 if target else None
    return da, dp


def main():
    DATA.mkdir(parents=True, exist_ok=True)
    DOCS.mkdir(parents=True, exist_ok=True)

    wb_f = load_workbook(ORIGINAL, data_only=False, read_only=True)
    wb_v = load_workbook(ORIGINAL, data_only=True, read_only=True)
    wsf = wb_f["Bajadas"]
    wsv = wb_v["Bajadas"]

    used_range = wsf.calculate_dimension(force=True)

    keywords = [
        "stickers", "sticker", "autoadhesivo", "adhesivo", "opp", "opp blanco", "blanco",
        "especial", "papel", "ilustración 90g", "ilustracion 90g", "obra 90g",
    ]
    keyword_hits = []
    for row in wsf.iter_rows(min_row=1, max_row=292, min_col=1, max_col=58):
        for c in row:
            v = c.value
            if isinstance(v, str):
                t = v.strip().lower()
                if any(k in t for k in keywords):
                    keyword_hits.append({"cell": c.coordinate, "value": v})

    ranges = [
        {"range": "B4:B10", "labels": [wsv[f"B{i}"].value for i in range(4, 11)]},
        {"range": "S4:S10", "title": wsf["S2"].value, "caras": wsf["S3"].value},
        {"range": "U4:U10", "title": wsf["U2"].value, "caras": wsf["U3"].value},
    ]

    formulas = []
    depend_sheets = set()
    for col, label in [("S", "papel_base_stickers"), ("U", "especial_base_opp_blanco")]:
        for r in range(4, 11):
            c = f"{col}{r}"
            f = wsf[c].value
            v = wsv[c].value
            refs = parse_refs(f)
            ref_nodes = []
            for sh, rc in refs:
                depend_sheets.add(sh)
                rv = wb_v[sh][rc].value if sh in wb_v.sheetnames else None
                rf = wb_f[sh][rc].value if sh in wb_f.sheetnames else None
                ref_nodes.append({"sheet": sh, "cell": rc, "value": rv, "formula": rf})
            formulas.append({
                "cell": c,
                "column_type": label,
                "range_label": wsv[f"B{r}"].value,
                "formula": f,
                "value": v,
                "references": ref_nodes,
            })

    historical_visible = []
    for r in range(4, 11):
        historical_visible.append({
            "range_label": wsv[f"B{r}"].value,
            "papel_visible_W": wsv[f"W{r}"].value,
            "especial_visible_X": wsv[f"X{r}"].value,
            "formula_W": wsf[f"W{r}"].value,
            "formula_X": wsf[f"X{r}"].value,
        })

    obj_rows, comparison = [], []
    for i, (rng, p_obj, e_obj) in enumerate(TARGETS, start=4):
        p_excel_formula = wsv[f"S{i}"].value
        e_excel_formula = wsv[f"U{i}"].value
        p_excel_visible = wsv[f"W{i}"].value
        e_excel_visible = wsv[f"X{i}"].value

        obj_rows.append({
            "categoria": "Bajadas Autoadhesivas",
            "modo_color": "fullcolor",
            "formato": "A3+",
            "tipo_producto": "autoadhesiva",
            "base_origen_excel": "stickers",
            "columna_precio": "papel",
            "material_excel": "stickers",
            "cantidad_rango": rng,
            "precio_objetivo_sin_iva": p_obj,
            "fuente_objetivo": "pdf_actualizado",
            "activo": True,
        })
        obj_rows.append({
            "categoria": "Bajadas Autoadhesivas",
            "modo_color": "fullcolor",
            "formato": "A3+",
            "tipo_producto": "autoadhesiva",
            "base_origen_excel": "OPP blanco",
            "columna_precio": "especial",
            "material_excel": "OPP blanco/blanco",
            "cantidad_rango": rng,
            "precio_objetivo_sin_iva": e_obj,
            "fuente_objetivo": "pdf_actualizado",
            "activo": True,
        })

        da_p_f, dp_p_f = calc_diff(float(p_excel_formula), p_obj)
        da_e_f, dp_e_f = calc_diff(float(e_excel_formula), e_obj)
        da_p_v, dp_p_v = calc_diff(float(p_excel_visible), p_obj)
        da_e_v, dp_e_v = calc_diff(float(e_excel_visible), e_obj)

        comparison.append({
            "cantidad_rango": rng,
            "papel": {
                "precio_excel_historico_formula_stickers": p_excel_formula,
                "precio_excel_historico_visible_W": p_excel_visible,
                "precio_objetivo_pdf": p_obj,
                "diff_formula_abs": da_p_f,
                "diff_formula_pct": dp_p_f,
                "diff_visible_abs": da_p_v,
                "diff_visible_pct": dp_p_v,
                "causa_probable": "dolar viejo / coeficiente viejo / pendiente interpretación de escala stickers",
            },
            "especial": {
                "precio_excel_historico_formula_opp": e_excel_formula,
                "precio_excel_historico_visible_X": e_excel_visible,
                "precio_objetivo_pdf": e_obj,
                "diff_formula_abs": da_e_f,
                "diff_formula_pct": dp_e_f,
                "diff_visible_abs": da_e_v,
                "diff_visible_pct": dp_e_v,
                "causa_probable": "dolar viejo / papel desactualizado / margen viejo",
            },
        })

    def node(sheet, cell, depth=0, maxd=2):
        f = wb_f[sheet][cell].value if sheet in wb_f.sheetnames else None
        v = wb_v[sheet][cell].value if sheet in wb_v.sheetnames else None
        n = {"sheet": sheet, "cell": cell, "value": v, "formula": f, "depth": depth, "children": []}
        if depth >= maxd:
            return n
        for sh, rc in parse_refs(f):
            if sh in wb_f.sheetnames:
                n["children"].append(node(sh, rc, depth + 1, maxd))
        return n

    trace = []
    for col, kind, origin in [("S", "papel", "stickers"), ("U", "especial", "OPP blanco/blanco")]:
        for r in range(4, 11):
            c = f"{col}{r}"
            trace.append({
                "categoria": "Bajadas Autoadhesivas",
                "formato": "A3+",
                "columna": kind,
                "base_origen_excel": origin,
                "range_label": wsv[f"B{r}"].value,
                "output_cell": c,
                "tree": node("Bajadas", c, 0, 2),
            })

    inventory = {
        "source_excel": str(ORIGINAL),
        "sheet": "Bajadas",
        "used_range": used_range,
        "keyword_hits": keyword_hits,
        "detected_ranges": ranges,
        "historical_visible_prices": historical_visible,
    }

    formula_out = {
        "formulas": formulas,
        "key_cells": {
            k: {"formula": wsf[k].value, "value": wsv[k].value}
            for k in ["S2", "U2", "S3", "U3", "AC12", "AC13", "AC48", "AH48", "AH49", "AH50", "AH51", "AH52", "AH53", "AH54", "AM38", "AM42"]
        },
    }

    dep_out = {
        "sheets_referenced": sorted(depend_sheets),
        "external_like_references": [],
        "notes": ["Dependencia principal detectada hacia hoja PAPEL US$."]
    }

    (DATA / "autoadhesivas_inventory.json").write_text(json.dumps(inventory, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA / "autoadhesivas_formulas.json").write_text(json.dumps(formula_out, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA / "autoadhesivas_dependencies.json").write_text(json.dumps(dep_out, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA / "autoadhesivas_price_trace_tree.json").write_text(json.dumps(trace, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA / "precios_autoadhesivas_objetivo.json").write_text(json.dumps(obj_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA / "autoadhesivas_excel_vs_objetivo.json").write_text(json.dumps(comparison, ensure_ascii=False, indent=2), encoding="utf-8")

    with (DATA / "precios_autoadhesivas_objetivo.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(obj_rows[0].keys()))
        w.writeheader()
        w.writerows(obj_rows)

    md1 = [
        "# 01 Diagnóstico Autoadhesivas", "",
        f"- Excel analizado (solo lectura): `{ORIGINAL}`",
        "- Hoja foco: `Bajadas`",
        f"- Rango usado detectado: `{used_range}`",
        "", "## Hallazgos clave",
        "- `Sticker` detectado en `S2` (columna autoadhesivas base papel).",
        "- `OPP  blco o blanco` detectado en `U2` (columna especial).",
        "- Escalas detectadas en `B4:B10`.",
        "- Fórmulas principales: `S4:S10` y `U4:U10`.",
        "- Dependencias a `PAPEL US$` vía celdas intermedias (por ejemplo `AC12`, `AC13`).",
        "", "## Rango autoadhesivas identificado",
        "- Encabezados: `S2:U3`",
        "- Fórmulas por rango: `S4:U10`",
        "- Valores visibles históricos (lista): `W4:X10` (sin fórmula en estas filas).",
    ]
    (DOCS / "01_diagnostico_autoadhesivas.md").write_text("\n".join(md1), encoding="utf-8")

    md2 = [
        "# 02 Árbol de Fórmula Autoadhesivas", "",
        "## Papel (base stickers)",
        "- Salida: `S4:S10`",
        "- Patrón: `=(AC4x + AC12*AF4x)*AH4x`",
        "- Lectura funcional: `(costo impresión/click + componente stickers ajustado) * coeficiente por escala`",
        "", "## Especial (base OPP blanco/blanco)",
        "- Salida: `U4:U10`",
        "- Patrón: `=(AC4x + AC13*AF4x)*AH4x`",
        "- Lectura funcional: `(costo impresión/click + componente OPP/blanco ajustado) * coeficiente por escala`",
        "", "## Celdas núcleo",
        "- `AC12`: componente stickers (deriva de PAPEL US$ + coeficientes)",
        "- `AC13`: componente OPP/blanco",
        "- `AC48:AC55`: costo base click/impresión por tramo",
        "- `AH48:AH55`: coeficientes por escala",
        "- `AF48:AF55`: factor multiplicador intermedio (actualmente 1.0)",
    ]
    (DOCS / "02_arbol_formula_autoadhesivas.md").write_text("\n".join(md2), encoding="utf-8")

    md3 = [
        "# 03 Origen Stickers y OPP", "",
        "## Stickers (columna papel)",
        "- Encabezado origen: `S2=Sticker`",
        "- Celdas salida: `S4:S10`",
        "- Fórmula usa explícitamente `AC12` como componente de stickers.",
        "- `AC12` se calcula con referencias a `PAPEL US$` y coeficientes (`AM38`, `AC26`).",
        "", "## OPP blanco/blanco (columna especial)",
        "- Encabezado origen: `U2=OPP  blco o blanco`",
        "- Celdas salida: `U4:U10`",
        "- Fórmula usa explícitamente `AC13` como componente de OPP/blanco.",
        "- `AC13` se calcula con referencias a `PAPEL US$` y coeficientes (`AM38`, `AC27`).",
        "", "## Comparación Excel histórico vs objetivo PDF", "",
        "| Rango | Papel Excel fórmula (S) | Papel Excel visible (W) | Papel Objetivo | Especial Excel fórmula (U) | Especial Excel visible (X) | Especial Objetivo |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    for c in comparison:
        md3.append(
            f"| {c['cantidad_rango']} | {c['papel']['precio_excel_historico_formula_stickers']:.2f} | {c['papel']['precio_excel_historico_visible_W']:.2f} | {c['papel']['precio_objetivo_pdf']:.2f} | {c['especial']['precio_excel_historico_formula_opp']:.2f} | {c['especial']['precio_excel_historico_visible_X']:.2f} | {c['especial']['precio_objetivo_pdf']:.2f} |"
        )
    (DOCS / "03_origen_stickers_y_opp.md").write_text("\n".join(md3), encoding="utf-8")

    md4 = [
        "# 04 Modelo Recomendado Autoadhesivas (base Excel)", "",
        "## Modelos evaluados",
        "- **Modelo A**: mantener fórmula Excel y actualizar solo dólar/materiales.",
        "- **Modelo B**: mantener estructura Excel + factor de calibración por columna (`papel`, `especial`).",
        "- **Modelo C**: mantener trazabilidad Excel pero resolver precio final por tabla calibrada por rango.",
        "", "## Recomendación",
        "- **Modelo recomendado: B (híbrido simple)**.",
        "", "Justificación:",
        "1. Se conserva la lógica estructural del Excel (`S4:S10`, `U4:U10`) y su trazabilidad.",
        "2. El componente `stickers` muestra magnitudes históricas no directamente compatibles con precio objetivo actual; requiere calibración explícita.",
        "3. `especial` (U) está en escala de magnitud cercana al objetivo y se beneficia de calibración moderada.",
        "4. Evita convertir todo en precio fijo, pero permite llegar al objetivo PDF actualizado por rango.",
        "", "## Pendientes de interpretación",
        "- Relación exacta entre salidas de fórmula `S/U` y valores visibles históricos `W/X`.",
        "- Confirmar si `W/X` fueron lista manual comercial o salida derivada externa.",
        "- Validar negocio sobre uso de `AM38`, `AM42` y política de margen actual.",
    ]
    (DOCS / "04_modelo_recomendado_autoadhesivas_base.md").write_text("\n".join(md4), encoding="utf-8")

    print("OK: análisis autoadhesivas generado")


if __name__ == "__main__":
    main()
