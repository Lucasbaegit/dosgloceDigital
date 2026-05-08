import ast
import json
import math
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from xl_byn_1_1_correction import apply_xl_byn_1_1_parenthesis_fix

BASE = Path(__file__).resolve().parents[2]
DATA_B2 = BASE / 'data' / 'bajadas_v2'
DATA_B = BASE / 'data' / 'bajadas'
DOCS_B2 = BASE / 'docs' / 'bajadas_v2'

EXCEL = BASE / 'original' / 'Copia de DIGITAL sistema 2025.xlsx'

OUT_AUDIT = DATA_B2 / 'auditoria_formula_xl_byn_1_1.json'
OUT_SIM = DATA_B2 / 'simulacion_correccion_xl_byn_1_1.json'
OUT_COMP = DATA_B2 / 'comparativa_xl_byn_1_1_corregido.json'
OUT_MD10 = DOCS_B2 / '10_auditoria_formula_xl_byn_1_1.md'
OUT_MD11 = DOCS_B2 / '11_simulacion_correccion_xl_byn_1_1.md'


def safe_eval(expr: str):
    node = ast.parse(expr, mode='eval')

    def ev(n):
        if isinstance(n, ast.Expression):
            return ev(n.body)
        if isinstance(n, ast.BinOp):
            l = ev(n.left)
            r = ev(n.right)
            if isinstance(n.op, ast.Add):
                return l + r
            if isinstance(n.op, ast.Sub):
                return l - r
            if isinstance(n.op, ast.Mult):
                return l * r
            if isinstance(n.op, ast.Div):
                return l / r if r != 0 else math.nan
            raise ValueError('op')
        if isinstance(n, ast.UnaryOp):
            v = ev(n.operand)
            if isinstance(n.op, ast.USub):
                return -v
            if isinstance(n.op, ast.UAdd):
                return +v
            raise ValueError('uop')
        if isinstance(n, ast.Constant):
            return float(n.value)
        raise ValueError('node')

    return float(ev(node))


def col_to_num(c):
    n = 0
    for ch in c:
        n = n * 26 + ord(ch) - 64
    return n


def get_val(ws, coord):
    v = ws[coord].value
    if isinstance(v, (int, float)):
        return float(v)
    return None


def eval_formula_with_values(formula, ws_values):
    if not formula or not formula.startswith('='):
        return None
    expr = formula[1:]
    for ref in sorted(set(re.findall(r'\$?[A-Z]{1,3}\$?\d+', expr)), key=len, reverse=True):
        coord = ref.replace('$', '')
        v = get_val(ws_values, coord)
        expr = expr.replace(ref, str(v if v is not None else 0))

    try:
        return safe_eval(expr)
    except Exception:
        return None


def estado(diff_pct):
    if diff_pct is None:
        return 'SIN_COMPARACION'
    d = abs(diff_pct)
    if d <= 2:
        return 'OK'
    if d <= 5:
        return 'DIFERENCIA_LEVE'
    if d <= 10:
        return 'DIFERENCIA_MEDIA'
    return 'DIFERENCIA_ALTA'


def main():
    comp_c = json.loads((DATA_B2 / 'comparativa_modelo_c.json').read_text(encoding='utf-8'))
    cfg = json.loads((DATA_B2 / 'bajadas_v2_config_recalibrada.json').read_text(encoding='utf-8'))
    validated = json.loads((DATA_B2 / 'precios_pdf_objetivo_validado.json').read_text(encoding='utf-8'))
    formulas = json.loads((DATA_B / 'bajadas_formulas.json').read_text(encoding='utf-8'))['formulas']

    wb_formula = load_workbook(EXCEL, data_only=False, read_only=True)
    wb_values = load_workbook(EXCEL, data_only=True, read_only=True)
    wsf = wb_formula['Bajadas']
    wsv = wb_values['Bajadas']

    formula_map = {f['cell']: f['formula'] for f in formulas}

    # XL ByN block mapping
    # row 63 materials, row64 faces, row65..70 qty
    row_material = 63
    row_face = 64
    qty_rows = list(range(65, 71))
    qty_map = {r: str(wsv[f'B{r}'].value).strip() if wsv[f'B{r}'].value is not None else None for r in qty_rows}

    # detect affected 1/1 cells in XL ByN formulas
    affected = []
    compare_refs = {'A3+': [], 'XA3': [], 'A4': [], 'XL': []}

    # helper to get matching columns by face row in a range
    def cells_for_face(rmin, rmax, face='1/1'):
        out = []
        for cnum in range(col_to_num('C'), col_to_num('T') + 1):
            col = ''
            n = cnum
            while n:
                n, rem = divmod(n - 1, 26)
                col = chr(65 + rem) + col
            if str(wsv[f'{col}{row_face}'].value).strip() == face:
                for r in range(rmin, rmax + 1):
                    cell = f'{col}{r}'
                    if cell in formula_map:
                        out.append(cell)
        return out

    # Reference structures A3+/XA3/A4/XL (ByN sections)
    compare_refs['A3+'] = cells_for_face(47, 53, '1/1')
    compare_refs['XA3'] = cells_for_face(56, 62, '1/1')
    compare_refs['A4'] = cells_for_face(74, 79, '1/1')
    compare_refs['XL'] = cells_for_face(65, 70, '1/1')

    # CSV objective lookup for XL ByN
    target_lookup = {}
    for r in comp_c:
        if r['categoria']=='Bajadas Blanco y Negro' and r['formato']=='XL' and r['caras']=='1/1':
            k = (r['tipo_papel'], r['material'], r['gramaje'], str(r['cantidad_rango']).strip(), r['caras'])
            target_lookup[k] = r

    # material map from row63 by column pairs
    material_cache = {}
    for c in range(col_to_num('C'), col_to_num('T') + 1):
        col=''; n=c
        while n:
            n, rem = divmod(n-1,26)
            col = chr(65+rem)+col
        material = wsv[f'{col}{row_material}'].value
        face = wsv[f'{col}{row_face}'].value
        material_cache[col] = (str(material).strip() if material is not None else None, str(face).strip() if face is not None else None)

    component_outside = []
    simulations = []

    for cell in compare_refs['XL']:
        col = re.match(r'([A-Z]+)', cell).group(1)
        row = int(re.findall(r'\d+', cell)[0])
        formula = formula_map.get(cell)
        if not formula:
            continue
        mat, face = material_cache[col]
        if face != '1/1':
            continue

        correction = apply_xl_byn_1_1_parenthesis_fix(
            formula,
            formato='XL',
            modo_color='blanco_y_negro',
            caras='1/1',
        )
        corrected = correction.formula_corregida
        normalized_expected = correction.formula_corregida
        if correction.applied:
            component_outside.append({
                'cell': cell,
                'component': 'AG48',
                'original': correction.formula_original_detectada,
                'corrected': correction.formula_corregida,
                'motivo': correction.motivo,
                'componente_corregido': correction.componente_corregido,
            })

        original_price = get_val(wsv, cell)
        sim_price = eval_formula_with_values(corrected, wsv)

        # infer tipo_papel by gramaje/material text
        mat_low = (mat or '').lower()
        if 'triplex' in mat_low or any(g in mat_low for g in ['350','300','200']):
            tipo_papel = 'pesado'
        else:
            tipo_papel = 'liviano'
        gram = mat if any(ch.isdigit() for ch in (mat or '')) else None

        key = (tipo_papel, mat if mat is not None else None, gram if gram is not None else None, qty_map[row], '1/1')
        t = target_lookup.get(key)
        # fallback looser match by qty+caras+tipo
        if t is None:
            cands = [x for x in comp_c if x['categoria']=='Bajadas Blanco y Negro' and x['formato']=='XL' and x['caras']=='1/1' and str(x['cantidad_rango']).strip()==str(qty_map[row]).strip() and x['tipo_papel']==tipo_papel]
            # choose by gramage text similarity
            if cands:
                t = cands[0]

        target = t['precio_objetivo_csv'] if t else None
        diff_orig = None if (target is None or original_price is None or target==0) else ((original_price-target)/target*100)
        diff_corr = None if (target is None or sim_price is None or target==0) else ((sim_price-target)/target*100)
        mejora = None
        if diff_orig is not None and diff_corr is not None:
            mejora = abs(diff_orig) - abs(diff_corr)

        simulations.append({
            'cell': cell,
            'cantidad_rango': qty_map[row],
            'material_detectado': mat,
            'tipo_papel_inferido': tipo_papel,
            'formula_original': correction.formula_original_detectada,
            'formula_normalizada_esperada': normalized_expected,
            'formula_simulada_corregida': correction.formula_corregida,
            'motivo_correccion': correction.motivo,
            'componente_corregido': correction.componente_corregido,
            'correccion_aplicada': correction.applied,
            'precio_original_excel': original_price,
            'precio_simulado_corregido': sim_price,
            'precio_objetivo_csv': target,
            'diferencia_original_pct': diff_orig,
            'diferencia_corregida_pct': diff_corr,
            'mejora_pct_abs': mejora,
        })

    # scenarios on XL ByN 1/1 segment
    seg_A = [x for x in comp_c if x['categoria']=='Bajadas Blanco y Negro' and x['formato']=='XL' and x['caras']=='1/1']
    # scenario B/C use corrected sim for matching cells, keep same factors (no new rule)
    map_sim = {x['cell']: x for x in simulations}

    # Need map from cell->row attributes for comparison; use simulations as audited set
    scen = []
    for s in simulations:
        if s['precio_objetivo_csv'] is None:
            stA = 'SIN_COMPARACION'; stB = 'SIN_COMPARACION';
        else:
            stA = estado(s['diferencia_original_pct'])
            stB = estado(s['diferencia_corregida_pct'])
        scen.append({
            'cell': s['cell'],
            'cantidad_rango': s['cantidad_rango'],
            'precio_objetivo_csv': s['precio_objetivo_csv'],
            'escenario_A_precio': s['precio_original_excel'],
            'escenario_A_diff_pct': s['diferencia_original_pct'],
            'escenario_A_estado': stA,
            'escenario_B_precio': s['precio_simulado_corregido'],
            'escenario_B_diff_pct': s['diferencia_corregida_pct'],
            'escenario_B_estado': stB,
            'escenario_C_precio': s['precio_simulado_corregido'],
            'escenario_C_diff_pct': s['diferencia_corregida_pct'],
            'escenario_C_estado': stB,
            'mejora_pct_abs': s['mejora_pct_abs'],
        })

    def metrics(items, prefix):
        diffs = [i[f'{prefix}_diff_pct'] for i in items if i[f'{prefix}_diff_pct'] is not None]
        cnt = Counter(i[f'{prefix}_estado'] for i in items)
        return {
            'OK': cnt.get('OK',0),
            'DIFERENCIA_LEVE': cnt.get('DIFERENCIA_LEVE',0),
            'DIFERENCIA_MEDIA': cnt.get('DIFERENCIA_MEDIA',0),
            'DIFERENCIA_ALTA': cnt.get('DIFERENCIA_ALTA',0),
            'SIN_COMPARACION': cnt.get('SIN_COMPARACION',0),
            'error_promedio_pct_abs': (sum(abs(x) for x in diffs)/len(diffs)) if diffs else None,
            'error_mediano_pct_abs': (sorted(abs(x) for x in diffs)[len(diffs)//2] if diffs else None),
            'error_max_pct_abs': (max(abs(x) for x in diffs) if diffs else None),
        }

    mA = metrics(scen, 'escenario_A')
    mB = metrics(scen, 'escenario_B')
    mC = metrics(scen, 'escenario_C')

    audit = {
        'no_excel_modificado': True,
        'segmento': 'Bajadas Blanco y Negro / XL / caras 1/1',
        'celdas_a3_byn_1_1': compare_refs['A3+'],
        'celdas_xa3_byn_1_1': compare_refs['XA3'],
        'celdas_a4_byn_1_1': compare_refs['A4'],
        'celdas_xl_byn_1_1': compare_refs['XL'],
        'componente_fuera_parentesis_detectado': 'AG48 multiplicando fuera del paréntesis en XL ByN 1/1',
        'ocurrencias_componente_fuera': component_outside,
        'comparacion_estructural': {
            'A3+_1_1_patron': '(variable*AG + papel) * AH',
            'XA3_1_1_patron': '(variable*AG + papel) * AH',
            'A4_1_1_patron': '(variable*AG + papel) * AH',
            'XL_1_1_patron_actual': '(variable + papel) * AH * AG',
            'XL_1_1_patron_esperado': '(variable*AG + papel) * AH',
        },
    }

    OUT_AUDIT.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding='utf-8')
    OUT_SIM.write_text(json.dumps(simulations, ensure_ascii=False, indent=2), encoding='utf-8')
    OUT_COMP.write_text(json.dumps({'escenarios':scen,'metricas':{'A':mA,'B':mB,'C':mC}}, ensure_ascii=False, indent=2), encoding='utf-8')

    md10 = []
    md10 += [
        '# 10 - Auditoría fórmula XL ByN 1/1',
        '',
        '- Se trabajó en modo simulación/auditoría. No se modificó el Excel.',
        '',
        '## Fórmula problemática',
        '- Patrón XL ByN 1/1 detectado: `(paper + variable) * AH * AG`',
        '- Patrón esperado (consistente con A3+/XA3/A4 1/1): `(paper + variable*AG) * AH`',
        '- Componente fuera del paréntesis: `AG48`.',
        '',
        '## Celdas XL ByN 1/1 afectadas',
        f"- {', '.join(compare_refs['XL'])}",
        '',
        '## Comparación estructural',
        '- A3+ 1/1: `(variable*AG + papel) * AH`',
        '- XA3 1/1: `(variable*AG + papel) * AH`',
        '- A4 1/1: `(variable*AG + papel) * AH`',
        '- XL 1/1 actual: `(variable + papel) * AH * AG`',
        '- XL 1/1 esperado: `(variable*AG + papel) * AH`',
    ]
    OUT_MD10.write_text('\n'.join(md10), encoding='utf-8')

    md11 = []
    md11 += [
        '# 11 - Simulación corrección XL ByN 1/1',
        '',
        f"- Casos simulados: {len(simulations)}",
        f"- Diferencias altas antes (A): {mA['DIFERENCIA_ALTA']}",
        f"- Diferencias altas después (B/C): {mB['DIFERENCIA_ALTA']}",
        f"- Mejora promedio (% abs): {round(sum(x['mejora_pct_abs'] for x in simulations if x['mejora_pct_abs'] is not None)/max(1,len([x for x in simulations if x['mejora_pct_abs'] is not None])),4) if simulations else 'N/A'}",
        '',
        '## Métricas por escenario',
        f"- Escenario A: {mA}",
        f"- Escenario B: {mB}",
        f"- Escenario C: {mC}",
        '',
        '## Recomendación preliminar',
        '- Si la corrección reduce materialmente DIFERENCIA_ALTA, aplicar corrección lógica en Bajadas v2 antes de crear regla especial.',
        '- Si no alcanza, recién ahí evaluar regla especial XL ByN.',
    ]
    OUT_MD11.write_text('\n'.join(md11), encoding='utf-8')

    mejora_vals = [x['mejora_pct_abs'] for x in simulations if x['mejora_pct_abs'] is not None]
    mejora_prom = (sum(mejora_vals)/len(mejora_vals)) if mejora_vals else None

    print(json.dumps({
        'no_excel_modificado': True,
        'xl_byn_1_1_cells': compare_refs['XL'],
        'outside_component': 'AG48',
        'comparative_cells': {
            'A3+': len(compare_refs['A3+']),
            'XA3': len(compare_refs['XA3']),
            'A4': len(compare_refs['A4']),
            'XL': len(compare_refs['XL']),
        },
        'cases_simulated': len(simulations),
        'altas_before': mA['DIFERENCIA_ALTA'],
        'altas_after': mB['DIFERENCIA_ALTA'],
        'avg_improvement_pct_abs': mejora_prom,
        'files': [str(OUT_AUDIT), str(OUT_SIM), str(OUT_COMP), str(OUT_MD10), str(OUT_MD11)]
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
