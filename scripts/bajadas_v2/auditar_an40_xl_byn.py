import json
import re
import statistics
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


def main() -> None:
    root = Path(__file__).resolve().parents[2]
    wb_path = root / "original" / "Copia de DIGITAL sistema 2025.xlsx"

    formulas_data = json.loads((root / "data/bajadas/bajadas_formulas.json").read_text(encoding="utf-8"))
    blocks_data = json.loads((root / "data/bajadas/bajadas_blocks_semantic.json").read_text(encoding="utf-8"))
    comp_d = json.loads((root / "data/bajadas_v2/comparativa_modelo_d_final.json").read_text(encoding="utf-8"))

    wb_formula = load_workbook(wb_path, data_only=False, read_only=True)
    wb_values = load_workbook(wb_path, data_only=True, read_only=True)
    wsf = wb_formula["Bajadas"]
    wsv = wb_values["Bajadas"]

    def in_range(cell: str, rng: str) -> bool:
        start, end = rng.split(":")
        c1, r1 = coordinate_from_string(start)
        c2, r2 = coordinate_from_string(end)
        col = column_index_from_string(re.match(r"[A-Z]+", cell).group(0))
        row = int(re.search(r"\d+", cell).group(0))
        return (
            column_index_from_string(c1) <= col <= column_index_from_string(c2)
            and r1 <= row <= r2
        )

    blocks = []
    for b in blocks_data.get("blocks", []):
        blocks.append((b["id"], b["range"], b.get("title_detected", ""), b.get("function_probable", "")))

    def block_of(cell: str):
        for bid, rng, title, func in blocks:
            if in_range(cell, rng):
                return {"id": bid, "range": rng, "title": title, "function": func}
        return None

    an40_formula = wsf["AN40"].value
    an40_value = wsv["AN40"].value

    neighbor_labels = []
    for rr in range(38, 43):
        for cc in ["AL", "AM", "AN", "AO", "AP"]:
            v = wsf[f"{cc}{rr}"].value
            if isinstance(v, str) and not v.startswith("=") and v.strip():
                neighbor_labels.append(f"{cc}{rr}:{v}")

    an40_block = block_of("AN40")

    formula_entries = formulas_data.get("formulas", [])
    an40_forms = [f for f in formula_entries if "AN40" in (f.get("formula") or "")]

    family_by_block = {
        "SB02": "A3+",
        "SB03": "XA3",
        "SB04": "XL 32x70",
        "SB06": "A4 color",
        "SB07": "A3 PLUS pack color",
        "SB08": "XA3 ByN",
        "SB09": "XL ByN",
        "SB11": "A4 ByN",
    }

    def mode_by_block(bid):
        return "blanco_y_negro" if bid in {"SB08", "SB09", "SB11"} else "fullcolor"

    ref_pat = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)")

    def get_value(ref):
        try:
            return float(wsv[ref].value)
        except Exception:
            return None

    an40_usage = []
    ratios_by_face = {"1/0": [], "1/1": []}

    for f in an40_forms:
        cell = f["cell"]
        formula = f["formula"]
        b = block_of(cell)
        bid = b["id"] if b else None
        mode = mode_by_block(bid) if bid else "pendiente"

        col_letters = re.match(r"[A-Z]+", cell).group(0)
        col_idx = column_index_from_string(col_letters)
        caras = "1/1" if (col_idx % 2 == 0) else "1/0"

        multiplied = None
        if "*AN40" in formula:
            m = re.search(r"([A-Z]{1,3}\d+)\*AN40", formula.replace("$", ""))
            multiplied = m.group(1) if m else "PENDIENTE"

        affects = "click" if multiplied and multiplied.startswith("AA") else "pendiente_de_interpretacion"

        paper = None
        click = None
        m2 = re.search(r"\(([^\+\)]+)\+([^\)]+)\*AN40\)", formula.replace("$", ""))
        if m2:
            paper = m2.group(1)
            click = m2.group(2)

        if paper and click:
            pv = get_value(paper)
            cv = get_value(click)
            av = float(an40_value) if an40_value is not None else None
            if pv is not None and cv is not None and av is not None and (pv + cv * av) != 0:
                ratios_by_face[caras].append((pv + cv) / (pv + cv * av))

        an40_usage.append(
            {
                "celda": cell,
                "bloque": bid,
                "familia_formato_probable": family_by_block.get(bid, "PENDIENTE_DE_INTERPRETACION"),
                "modo_color_probable": mode,
                "caras_probable": caras,
                "formula_original": formula,
                "componente_multiplicado_por_AN40": multiplied or "PENDIENTE_DE_INTERPRETACION",
                "an40_afecta": affects,
                "valor_actual_formula": wsv[cell].value,
                "salida_asociada": cell,
            }
        )

    wanted = {
        "XL_ByN_1_0": "=(AA11+AA48*AN40)*AH$48",
        "XL_Color_4_0": "=(AD48+AA11)*$AH$48",
    }
    found_struct = {}
    for key, val in wanted.items():
        hit = next((x for x in formula_entries if x.get("formula") == val), None)
        found_struct[key] = {"formula": val, "celda_ejemplo": hit.get("cell") if hit else None}

    uses_blocks = sorted({u["bloque"] for u in an40_usage if u["bloque"]})
    uses_modes = sorted({u["modo_color_probable"] for u in an40_usage})
    uses_faces = sorted({u["caras_probable"] for u in an40_usage})

    xl_byn = [
        r
        for r in comp_d["comparativa"]
        if r.get("categoria") == "Bajadas Blanco y Negro" and r.get("formato") == "XL"
    ]

    r10 = statistics.mean(ratios_by_face["1/0"]) if ratios_by_face["1/0"] else 1.0
    r11 = statistics.mean(ratios_by_face["1/1"]) if ratios_by_face["1/1"] else 1.0
    an40 = float(an40_value) if an40_value is not None else 2.0

    comparables = [
        r
        for r in xl_byn
        if r.get("estado") != "SIN_COMPARACION" and r.get("precio_objetivo_csv") not in (None, 0)
    ]

    best = None
    for k_int in range(50, 401):
        k = k_int / 100

        def ratio_from_r(rm):
            if abs(rm - 1) < 1e-9:
                return 1.0
            q = (1 - rm * an40) / (rm - 1)
            den = q + an40
            return (q + k) / den if den != 0 else 1.0

        rk10 = ratio_from_r(r10)
        rk11 = ratio_from_r(r11)

        errs = []
        for rr in comparables:
            base = float(rr["precio_estimado_v2"])
            ratio = rk11 if rr.get("caras") == "1/1" else rk10
            est = base * ratio
            obj = float(rr["precio_objetivo_csv"])
            errs.append(abs((est - obj) / obj) * 100)

        mae = sum(errs) / len(errs) if errs else 1e9
        if best is None or mae < best[0]:
            best = (mae, k, rk10, rk11)

    best_mae, best_k, best_r10, best_r11 = best

    def estimate(base, caras, scenario):
        if scenario == "A":
            return base
        if scenario == "B":
            return base * (r11 if caras == "1/1" else r10)
        if scenario == "C":
            return base if caras == "1/1" else base * r10
        if scenario == "D":
            return base * (best_r11 if caras == "1/1" else best_r10)
        return base

    def state(diff_pct):
        a = abs(diff_pct)
        if a <= 2:
            return "OK"
        if a <= 5:
            return "DIFERENCIA_LEVE"
        if a <= 10:
            return "DIFERENCIA_MEDIA"
        return "DIFERENCIA_ALTA"

    sim = {}
    for sc in ["A", "B", "C", "D"]:
        rows = []
        for rr in xl_byn:
            base = float(rr["precio_estimado_v2"]) if rr.get("precio_estimado_v2") is not None else None
            obj = rr.get("precio_objetivo_csv")

            if base is None:
                est = None
                st = "SIN_COMPARACION"
                dp = None
            elif rr.get("estado") == "SIN_COMPARACION" or obj in (None, 0):
                est = estimate(base, rr.get("caras"), sc)
                st = "SIN_COMPARACION"
                dp = None
            else:
                est = estimate(base, rr.get("caras"), sc)
                dp = ((est - float(obj)) / float(obj)) * 100
                st = state(dp)

            rows.append(
                {
                    "categoria": rr.get("categoria"),
                    "formato": rr.get("formato"),
                    "tipo_papel": rr.get("tipo_papel"),
                    "material": rr.get("material"),
                    "gramaje": rr.get("gramaje"),
                    "cantidad_rango": rr.get("cantidad_rango"),
                    "caras": rr.get("caras"),
                    "precio_objetivo_csv": obj,
                    "precio_estimado": est,
                    "diferencia_porcentual": dp,
                    "estado": st,
                }
            )

        comp = [r for r in rows if r["estado"] != "SIN_COMPARACION" and r["diferencia_porcentual"] is not None]
        errs = [abs(r["diferencia_porcentual"]) for r in comp]
        metric = {k: sum(1 for r in rows if r["estado"] == k) for k in ["OK", "DIFERENCIA_LEVE", "DIFERENCIA_MEDIA", "DIFERENCIA_ALTA", "SIN_COMPARACION"]}
        metric.update(
            {
                "error_promedio_abs": sum(errs) / len(errs) if errs else None,
                "error_mediano_abs": statistics.median(errs) if errs else None,
                "diferencia_maxima_abs": max(errs) if errs else None,
            }
        )
        sim[sc] = {"rows": rows, "metricas": metric}

    audit = {
        "no_excel_modificado": True,
        "no_pdf_ocr": True,
        "an40": {
            "celda": "AN40",
            "valor_actual": an40_value,
            "formula": an40_formula,
            "etiquetas_cercanas": neighbor_labels,
            "bloque_semantico": an40_block,
            "funcion_probable": "Multiplicador auxiliar asociado a costo variable/click en ByN (PENDIENTE_DE_INTERPRETACION).",
        },
        "uso_an40": {
            "cantidad_formulas": len(an40_forms),
            "formula_cells": [u["celda"] for u in an40_usage],
            "detalles": an40_usage,
            "segmentos_afectados": {
                "bloques": uses_blocks,
                "modos": uses_modes,
                "caras": uses_faces,
                "solo_xl_byn": uses_blocks == ["SB09"],
            },
        },
        "comparacion_estructural": {
            "xl_byn_1_0": found_struct.get("XL_ByN_1_0"),
            "xl_color_4_0": found_struct.get("XL_Color_4_0"),
            "lectura": "En XL ByN el AN40 multiplica el componente de click (AA48*AN40) antes de margen; en XL color no hay multiplicador equivalente en click dentro de la misma estructura.",
        },
    }

    sim_out = {
        "supuestos": {
            "an40_valor": an40,
            "ratio_remove_an40_por_cara": {"1/0": r10, "1/1": r11},
            "factor_calibrado_escenario_d": best_k,
            "ratio_factor_calibrado_por_cara": {"1/0": best_r10, "1/1": best_r11},
        },
        "escenarios": {
            "A_actual_con_AN40": sim["A"]["metricas"],
            "B_sin_AN40_click": sim["B"]["metricas"],
            "C_AN40_solo_1_1": sim["C"]["metricas"],
            "D_AN40_por_factor_calibrado": sim["D"]["metricas"],
        },
        "detalle_xl_byn": {k: v["rows"] for k, v in sim.items()},
    }

    (root / "data/bajadas_v2/auditoria_an40_xl_byn.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    (root / "data/bajadas_v2/simulacion_an40_xl_byn.json").write_text(json.dumps(sim_out, ensure_ascii=False, indent=2), encoding="utf-8")

    md1 = []
    md1.append("# 17 - Auditoría AN40 en XL ByN\n")
    md1.append(f"- AN40 valor actual: **{an40_value}**")
    md1.append(f"- AN40 fórmula: **{an40_formula}**")
    md1.append(f"- Bloque: **{an40_block['id'] if an40_block else 'N/A'}**")
    md1.append(f"- Fórmulas que usan AN40: **{len(an40_forms)}**")
    md1.append(f"- Bloques afectados: **{', '.join(uses_blocks)}**")
    md1.append(f"- Modos afectados: **{', '.join(uses_modes)}**")
    md1.append(f"- Caras afectadas: **{', '.join(uses_faces)}**")
    md1.append("\n## Etiquetas cercanas AN40")
    for e in neighbor_labels:
        md1.append(f"- {e}")
    md1.append("\n## Diferencia estructural ByN vs Fullcolor")
    md1.append("- XL ByN 1/0: `=(AA11+AA48*AN40)*AH$48`")
    md1.append("- XL Fullcolor 4/0: `=(AD48+AA11)*$AH$48`")
    md1.append("- Lectura: AN40 entra como multiplicador del click en ByN y no tiene homólogo directo en fullcolor dentro de la misma composición.")
    (root / "docs/bajadas_v2/17_auditoria_an40_xl_byn.md").write_text("\n".join(md1), encoding="utf-8")

    md2 = []
    md2.append("# 18 - Simulación AN40 en XL ByN\n")
    map_titles = [
        ("A_actual_con_AN40", "Escenario A (actual)"),
        ("B_sin_AN40_click", "Escenario B (sin AN40 en click)"),
        ("C_AN40_solo_1_1", "Escenario C (AN40 solo en 1/1)"),
        ("D_AN40_por_factor_calibrado", "Escenario D (AN40 reemplazado por factor calibrado)"),
    ]
    for key, title in map_titles:
        m = sim_out["escenarios"][key]
        md2.append(f"## {title}")
        md2.append(f"- OK: {m['OK']}")
        md2.append(f"- DIFERENCIA_LEVE: {m['DIFERENCIA_LEVE']}")
        md2.append(f"- DIFERENCIA_MEDIA: {m['DIFERENCIA_MEDIA']}")
        md2.append(f"- DIFERENCIA_ALTA: {m['DIFERENCIA_ALTA']}")
        md2.append(f"- SIN_COMPARACION: {m['SIN_COMPARACION']}")
        md2.append(f"- Error promedio abs: {m['error_promedio_abs']:.6f}")
        md2.append(f"- Error mediano abs: {m['error_mediano_abs']:.6f}")
        md2.append(f"- Diferencia máxima abs: {m['diferencia_maxima_abs']:.6f}\n")

    order = [("A", sim["A"]["metricas"]), ("B", sim["B"]["metricas"]), ("C", sim["C"]["metricas"]), ("D", sim["D"]["metricas"])]
    best_sc = sorted(order, key=lambda x: (x[1]["DIFERENCIA_ALTA"], x[1]["error_promedio_abs"]))[0][0]
    md2.append(f"## Mejor escenario por métricas\n- **{best_sc}**")
    md2.append("- Nota: esta simulación es de auditoría sobre el segmento XL ByN y no modifica configuración.")
    (root / "docs/bajadas_v2/18_simulacion_an40_xl_byn.md").write_text("\n".join(md2), encoding="utf-8")

    print("OK")


if __name__ == "__main__":
    main()
