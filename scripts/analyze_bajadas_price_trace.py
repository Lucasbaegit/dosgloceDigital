import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

BASE = Path(__file__).resolve().parents[1]
EXCEL = BASE / "original" / "Copia de DIGITAL sistema 2025.xlsx"
DATA = BASE / "data" / "bajadas"
DOCS = BASE / "docs" / "bajadas"

TRACE_JSON = DATA / "bajadas_price_trace_tree.json"
ALERTS_JSON = DATA / "bajadas_price_trace_alerts.json"
DOC_TRACE = DOCS / "06_arbol_trazabilidad_precios.md"
DOC_ALERTS = DOCS / "07_incongruencias_detectadas.md"
MMD = DOCS / "arbol_trazabilidad_bajadas.mmd"

MAX_DEPTH = 5
MAX_NODES_PER_TRACE = 120

TOKEN_TYPES = [
    ("precio_final", ["precio final", "final"]),
    ("precio_unitario", ["unit", "unitario"]),
    ("precio_base", ["base", "costo de bajada"]),
    ("coste_papel", ["papel", "pliego"]),
    ("coste_impresion", ["clic", "impres", "toner", "color", "byn", "b&n"]),
    ("coeficiente", ["coef"]),
    ("margen", ["ganan", "margen"]),
    ("cantidad", ["cantidad", " a "]),
    ("gramaje", ["g", "grs", "gram"]),
    ("formato", ["a3", "a4", "xa3", "xl", "cm", "x"]),
    ("caras_impresion", ["4/0", "4/4", "1/0", "1/1"]),
]

FAMILY_MAP = {
    "SB02": "A3+",
    "SB03": "XA3",
    "SB04": "XL 32x70",
    "SB06": "A4 color",
    "SB07": "A3 PLUS pack color",
    "SB08": "XA3 ByN",
    "SB09": "XL ByN",
    "SB10": "A3 PLUS pack color",
    "SB11": "A4 ByN",
}


def parse_range(rng: str):
    a, b = rng.split(":")
    m1 = re.match(r"([A-Z]+)(\d+)", a)
    m2 = re.match(r"([A-Z]+)(\d+)", b)
    c1, r1 = column_index_from_string(m1.group(1)), int(m1.group(2))
    c2, r2 = column_index_from_string(m2.group(1)), int(m2.group(2))
    return r1, c1, r2, c2


def coord_to_rc(coord: str):
    m = re.match(r"([A-Z]+)(\d+)", coord)
    return int(m.group(2)), column_index_from_string(m.group(1))


def rc_to_coord(r: int, c: int):
    return f"{get_column_letter(c)}{r}"


def refs_from_formula(formula: str):
    refs = []
    # cross-sheet refs
    for m in re.finditer(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_\. ]*))!\$?([A-Z]{1,3})\$?(\d+)", formula):
        sheet = (m.group(1) or m.group(2)).strip()
        coord = f"{m.group(3)}{m.group(4)}"
        refs.append((sheet, coord))

    # same-sheet refs (avoid those already captured with !)
    cleaned = re.sub(r"(?:'[^']+'|[A-Za-z_][A-Za-z0-9_\. ]*)!", "", formula)
    for m in re.finditer(r"\$?([A-Z]{1,3})\$?(\d+)", cleaned):
        refs.append((None, f"{m.group(1)}{m.group(2)}"))
    return refs


def value_kind(v):
    if v is None:
        return "empty"
    if isinstance(v, str):
        if v.startswith("="):
            return "formula"
        return "text"
    if isinstance(v, (int, float)):
        return "number"
    return "other"


def cell_label_context(ws, row, col):
    labels = []
    for rr in range(max(1, row - 1), row + 1):
        for cc in range(max(1, col - 3), col + 1):
            if rr == row and cc == col:
                continue
            v = ws.cell(rr, cc).value
            if isinstance(v, str) and v.strip() and not v.startswith("="):
                labels.append(v.strip())
    return labels[:5]


def probable_type_from_context(labels, formula=None, val=None):
    txt = " ".join(labels).lower()
    if formula:
        txt = (txt + " " + formula.lower()).strip()
    if isinstance(val, str):
        txt = (txt + " " + val.lower()).strip()

    for t, keys in TOKEN_TYPES:
        if any(k in txt for k in keys):
            return t

    if isinstance(val, (int, float)):
        return "auxiliar"
    return "pendiente_de_interpretacion"


def formula_incongruences(formula: str, refs: list):
    issues = []
    f = formula.replace(" ", "").upper()
    if "*1" in f or "1*" in f:
        issues.append(("multiplicacion_por_1", "leve", "La fórmula contiene multiplicación por 1."))
    if "/1" in f:
        issues.append(("division_por_1", "leve", "La fórmula contiene división por 1."))
    if "+0" in f or "-0" in f:
        issues.append(("suma_resta_0", "leve", "La fórmula contiene suma/resta de 0."))

    ref_only = [r[1] for r in refs if r[1]]
    dups = [x for x, c in Counter(ref_only).items() if c > 1]
    if dups:
        issues.append(("duplicacion_de_terminos", "media", f"Se repiten referencias en fórmula: {', '.join(dups[:5])}"))

    return issues


def build_trace(wb, ws_bajadas, block_by_cell, sem_blocks, start_cell, family, block_id):
    tree_nodes = []
    edges = []
    visited = set()
    local_alerts = []

    def walk(sheet_name, coord, depth, parent=None):
        key = (sheet_name, coord)
        if key in visited or len(tree_nodes) >= MAX_NODES_PER_TRACE or depth > MAX_DEPTH:
            return
        visited.add(key)

        ws = wb[sheet_name]
        row, col = coord_to_rc(coord)
        val = ws.cell(row, col).value
        formula = val if isinstance(val, str) and val.startswith("=") else None

        labels = cell_label_context(ws, row, col)
        b = block_by_cell.get(coord) if sheet_name == "Bajadas" else None
        ptype = probable_type_from_context(labels, formula=formula, val=None if formula else val)

        node = {
            "hoja": sheet_name,
            "celda": coord,
            "valor_actual": val,
            "formula": formula,
            "encabezados_cercanos": labels,
            "bloque_semantico": b,
            "tipo_probable": ptype,
            "depth": depth,
            "explicacion": "Nodo en cadena de cálculo/precio." if formula else "Dato/constante referenciado por cálculo.",
        }
        tree_nodes.append(node)
        if parent:
            edges.append({"from": parent, "to": f"{sheet_name}!{coord}"})

        if formula:
            refs = refs_from_formula(formula)
            local_alerts.extend([{
                "tipo": t,
                "severidad": s,
                "explicacion": e,
                "celda": f"{sheet_name}!{coord}",
            } for t, s, e in formula_incongruences(formula, refs)])

            for ref_sheet, ref_coord in refs:
                rs = ref_sheet if ref_sheet else sheet_name
                if rs not in wb.sheetnames:
                    local_alerts.append({
                        "tipo": "dependencia_externa_no_interpretada",
                        "severidad": "alta",
                        "explicacion": f"Referencia a hoja no disponible: {rs}!{ref_coord}",
                        "celda": f"{sheet_name}!{coord}",
                    })
                    continue

                if rs == "PAPEL US$" or rs == "Bajadas":
                    walk(rs, ref_coord, depth + 1, parent=f"{sheet_name}!{coord}")
                else:
                    # no explorar otras hojas, solo registrar
                    local_alerts.append({
                        "tipo": "dependencia_externa_no_interpretada",
                        "severidad": "media",
                        "explicacion": f"Referencia fuera de alcance: {rs}!{ref_coord}",
                        "celda": f"{sheet_name}!{coord}",
                    })

    walk("Bajadas", start_cell, 0, None)

    # post alerts on nodes
    by_key = {(n["hoja"], n["celda"]): n for n in tree_nodes}
    for n in tree_nodes:
        v = n["valor_actual"]
        if n["formula"]:
            continue
        if value_kind(v) == "empty":
            local_alerts.append({
                "tipo": "referencia_a_celda_vacia",
                "severidad": "media",
                "explicacion": "La cadena de cálculo usa una celda vacía.",
                "celda": f"{n['hoja']}!{n['celda']}",
            })
        if value_kind(v) == "text" and n["tipo_probable"] not in ("gramaje", "formato", "caras_impresion", "cantidad"):
            local_alerts.append({
                "tipo": "referencia_a_texto_inesperado",
                "severidad": "media",
                "explicacion": "Se referencia texto en cadena numérica potencial.",
                "celda": f"{n['hoja']}!{n['celda']}",
            })
        if value_kind(v) == "number" and not n["encabezados_cercanos"]:
            local_alerts.append({
                "tipo": "constante_sin_etiqueta_cercana",
                "severidad": "leve",
                "explicacion": "Constante numérica sin etiqueta cercana clara.",
                "celda": f"{n['hoja']}!{n['celda']}",
            })

    # decisions heuristic from formula at root
    root = next((n for n in tree_nodes if n["hoja"] == "Bajadas" and n["celda"] == start_cell), None)
    decisions = []
    if root and root.get("formula"):
        fr = root["formula"].lower()
        if "*" in fr:
            decisions.append({
                "descripcion": "Aplicación de coeficientes/multiplicadores en precio.",
                "celda_formula": f"Bajadas!{start_cell}",
                "justificacion_probable": "Ajuste por formato/rango/caras o margen.",
                "estado": "dudosa",
            })
        if "+" in fr:
            decisions.append({
                "descripcion": "Composición aditiva de costos parciales.",
                "celda_formula": f"Bajadas!{start_cell}",
                "justificacion_probable": "Suma de componentes (papel + impresión + ajustes).",
                "estado": "dudosa",
            })
        if "papel us$" in fr:
            decisions.append({
                "descripcion": "Consulta de costo externo de papel.",
                "celda_formula": f"Bajadas!{start_cell}",
                "justificacion_probable": "Vinculación a tabla de costo de papel.",
                "estado": "pendiente",
            })

    # missing explanations
    has_paper = any(n["tipo_probable"] == "coste_papel" or n["hoja"] == "PAPEL US$" for n in tree_nodes)
    has_print = any(n["tipo_probable"] == "coste_impresion" for n in tree_nodes)
    has_margin = any(n["tipo_probable"] in ("margen", "coeficiente") for n in tree_nodes)

    if not has_paper:
        local_alerts.append({
            "tipo": "salida_sin_explicacion_coste_papel",
            "severidad": "alta",
            "explicacion": "No se detectó nodo claro de coste de papel en la traza.",
            "celda": f"Bajadas!{start_cell}",
        })
    if not has_print:
        local_alerts.append({
            "tipo": "salida_sin_explicacion_coste_impresion",
            "severidad": "alta",
            "explicacion": "No se detectó nodo claro de coste de impresión en la traza.",
            "celda": f"Bajadas!{start_cell}",
        })
    if not has_margin:
        local_alerts.append({
            "tipo": "salida_sin_margen_claro",
            "severidad": "media",
            "explicacion": "No se detectó margen/coeficiente explícito en la traza.",
            "celda": f"Bajadas!{start_cell}",
        })

    return {
        "familia": family,
        "bloque": block_id,
        "celda_salida": start_cell,
        "encabezado_detectado": ", ".join(root["encabezados_cercanos"]) if root else "",
        "valor_actual": root["valor_actual"] if root else None,
        "formula": root["formula"] if root else None,
        "interpretacion_probable": "Precio calculado en tabla de Bajadas." if root and root.get("formula") else "PENDIENTE_DE_INTERPRETACION",
        "arbol_de_origen": tree_nodes,
        "aristas": edges,
        "decisiones_detectadas": decisions,
        "incongruencias": local_alerts,
    }


def main():
    DATA.mkdir(parents=True, exist_ok=True)
    DOCS.mkdir(parents=True, exist_ok=True)

    sem = json.loads((DATA / "bajadas_blocks_semantic.json").read_text(encoding="utf-8"))
    formulas = json.loads((DATA / "bajadas_formulas.json").read_text(encoding="utf-8"))

    wb = load_workbook(EXCEL, data_only=False, read_only=True)
    ws = wb["Bajadas"]

    sb_blocks = [b for b in sem["blocks"] if b["id"] in FAMILY_MAP]

    # map cell->block
    block_by_cell = {}
    for b in sem["blocks"]:
        r1, c1, r2, c2 = parse_range(b["range"])
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                block_by_cell[rc_to_coord(rr, cc)] = b["id"]

    # gather formula cells by block
    formulas_by_cell = {f["cell"]: f for f in formulas["formulas"]}
    block_formula_cells = defaultdict(list)
    block_referenced = defaultdict(set)

    for b in sb_blocks:
        r1, c1, r2, c2 = parse_range(b["range"])
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                coord = rc_to_coord(rr, cc)
                fo = formulas_by_cell.get(coord)
                if fo:
                    block_formula_cells[b["id"]].append(coord)
                    refs = refs_from_formula(fo["formula"])
                    for s, cref in refs:
                        if s is None and cref in block_by_cell and block_by_cell.get(cref) == b["id"]:
                            block_referenced[b["id"]].add(cref)

    # output candidate selection: not referenced inside block OR rightmost quantile columns
    traces = []
    all_alerts = []
    family_index = defaultdict(lambda: {"precios_trazados": 0, "leve": 0, "media": 0, "alta": 0})

    resolved_papel_refs = set()

    for b in sb_blocks:
        block_id = b["id"]
        fam = FAMILY_MAP[block_id]
        cells = block_formula_cells[block_id]
        if not cells:
            continue

        cols = sorted(coord_to_rc(c)[1] for c in cells)
        q_idx = int(len(cols) * 0.75)
        q_col = cols[q_idx] if cols else 1

        cands = []
        for c in cells:
            r, col = coord_to_rc(c)
            if c not in block_referenced[block_id] or col >= q_col:
                cands.append(c)

        # cap per block for usability
        cands = sorted(set(cands), key=lambda x: coord_to_rc(x))[:80]

        # incoherent jumps between adjacent columns by row among candidates with numeric values
        by_row = defaultdict(list)
        for c in cands:
            r, col = coord_to_rc(c)
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                by_row[r].append((col, c, float(v)))
        for r, vals in by_row.items():
            vals.sort()
            for i in range(1, len(vals)):
                prev = vals[i - 1]
                cur = vals[i]
                if prev[2] != 0:
                    ratio = cur[2] / prev[2]
                    if ratio > 2.5 or ratio < 0.4:
                        all_alerts.append({
                            "familia": fam,
                            "bloque": block_id,
                            "celda_salida": cur[1],
                            "tipo": "saltos_precio_incoherentes_columnas_vecinas",
                            "severidad": "media",
                            "explicacion": f"Salto potencialmente incoherente en fila {r}: {prev[1]}={prev[2]} -> {cur[1]}={cur[2]} (ratio {ratio:.2f}).",
                            "recomendacion": "Validar lógica de formato/gramaje/caras que justifica el salto.",
                        })

        for out_cell in cands:
            trace = build_trace(wb, ws, block_by_cell, sem["blocks"], out_cell, fam, block_id)
            traces.append(trace)
            family_index[fam]["precios_trazados"] += 1

            for n in trace["arbol_de_origen"]:
                if n["hoja"] == "PAPEL US$":
                    resolved_papel_refs.add(f"{n['hoja']}!{n['celda']}")

            for a in trace["incongruencias"]:
                record = {
                    "familia": fam,
                    "bloque": block_id,
                    "celda_salida": out_cell,
                    "tipo": a["tipo"],
                    "severidad": a["severidad"],
                    "explicacion": a["explicacion"],
                    "recomendacion": "Revisar la fórmula y confirmar intención de negocio.",
                    "celda_origen_alerta": a.get("celda"),
                }
                all_alerts.append(record)
                family_index[fam][a["severidad"]] += 1

    # top 20 by severity then frequency
    sev_rank = {"alta": 3, "media": 2, "leve": 1}
    all_alerts_sorted = sorted(all_alerts, key=lambda x: (sev_rank.get(x["severidad"], 0), x["familia"], x["tipo"]), reverse=True)

    trace_out = {
        "sheet": "Bajadas",
        "familias_cubiertas": sorted(set(FAMILY_MAP.values())),
        "precio_traces": traces,
        "total_traces": len(traces),
        "arboles_generados": len(traces),
        "referencias_papel_resueltas": sorted(resolved_papel_refs),
        "status": "PENDIENTE_DE_INTERPRETACION",
    }

    alerts_out = {
        "sheet": "Bajadas",
        "total_alerts": len(all_alerts),
        "top20_alerts": all_alerts_sorted[:20],
        "alerts": all_alerts_sorted,
        "family_summary": [{"familia": f, **vals} for f, vals in sorted(family_index.items())],
        "status": "PENDIENTE_DE_INTERPRETACION",
    }

    TRACE_JSON.write_text(json.dumps(trace_out, ensure_ascii=False, indent=2), encoding="utf-8")
    ALERTS_JSON.write_text(json.dumps(alerts_out, ensure_ascii=False, indent=2), encoding="utf-8")

    # Markdown trace doc
    md = []
    md.append("# 06 - Árbol de Trazabilidad de Precios (Bajadas)")
    md.append("")
    md.append("Estado: PENDIENTE_DE_INTERPRETACION")
    md.append("")
    md.append(f"- Celdas de salida trazadas: `{len(traces)}`")
    md.append(f"- Árboles generados: `{len(traces)}`")
    md.append(f"- Familias cubiertas: `{', '.join(sorted(set(FAMILY_MAP.values())))}`")
    md.append(f"- Referencias a PAPEL US$ resueltas: `{len(resolved_papel_refs)}`")
    md.append("")
    md.append("## Índice por familia")
    fam_groups = defaultdict(list)
    for t in traces:
        fam_groups[t["familia"]].append(t)
    for fam in sorted(fam_groups):
        md.append(f"### {fam}")
        for t in fam_groups[fam][:12]:
            md.append(f"- `{t['bloque']} / {t['celda_salida']}` valor=`{t['valor_actual']}` formula=`{t['formula']}`")

    md.append("")
    md.append("## Muestra de trazas (formato completo)")
    for t in traces[:12]:
        md.append("")
        md.append("Precio:")
        md.append(f"- familia: {t['familia']}")
        md.append(f"- bloque: {t['bloque']}")
        md.append(f"- celda_salida: {t['celda_salida']}")
        md.append(f"- encabezado_detectado: {t['encabezado_detectado']}")
        md.append(f"- valor_actual: {t['valor_actual']}")
        md.append(f"- fórmula: {t['formula']}")
        md.append(f"- interpretación_probable: {t['interpretacion_probable']}")
        md.append("- árbol_de_origen:")
        for i, n in enumerate(t["arbol_de_origen"][:10], 1):
            md.append(f"  - nodo {i}: hoja={n['hoja']} celda={n['celda']} valor={n['valor_actual']} tipo={n['tipo_probable']}")
        md.append("- decisiones_detectadas:")
        if t["decisiones_detectadas"]:
            for d in t["decisiones_detectadas"]:
                md.append(f"  - {d['descripcion']} | {d['celda_formula']} | estado={d['estado']}")
        else:
            md.append("  - PENDIENTE_DE_INTERPRETACION")
        md.append("- incongruencias:")
        if t["incongruencias"]:
            for a in t["incongruencias"][:5]:
                md.append(f"  - {a['tipo']} ({a['severidad']}): {a['explicacion']}")
        else:
            md.append("  - sin incongruencias detectadas en esta traza")

    DOC_TRACE.write_text("\n".join(md), encoding="utf-8")

    # Markdown alerts doc
    ad = []
    ad.append("# 07 - Incongruencias Detectadas")
    ad.append("")
    ad.append("Estado: PENDIENTE_DE_INTERPRETACION")
    ad.append("")
    ad.append(f"- Total alertas: `{len(all_alerts)}`")
    ad.append("")
    ad.append("## Top 20")
    for i, a in enumerate(all_alerts_sorted[:20], 1):
        ad.append(f"{i}. [{a['severidad'].upper()}] {a['familia']} {a['bloque']} {a['celda_salida']} -> {a['tipo']} | {a['explicacion']}")

    ad.append("")
    ad.append("## Resumen por familia")
    ad.append("| Familia | Precios trazados | Alertas leves | Alertas medias | Alertas altas |")
    ad.append("|---|---:|---:|---:|---:|")
    for fam, vals in sorted(family_index.items()):
        ad.append(f"| {fam} | {vals['precios_trazados']} | {vals['leve']} | {vals['media']} | {vals['alta']} |")

    DOC_ALERTS.write_text("\n".join(ad), encoding="utf-8")

    # Mermaid
    mm = []
    mm.append("flowchart TD")
    mm.append('  A["Bajadas"]')
    mm.append('  P["PAPEL US$"]')
    families = sorted(set(FAMILY_MAP.values()))
    for i, fam in enumerate(families, 1):
        fid = f"F{i}"
        mm.append(f'  {fid}["{fam}"]')
        mm.append(f"  A --> {fid}")

    # sample output nodes for visual clarity
    for idx, t in enumerate(traces[:36], 1):
        nid = f"O{idx}"
        label = f"{t['bloque']} {t['celda_salida']}"
        fam_id = f"F{families.index(t['familia']) + 1}"
        mm.append(f'  {nid}["{label}"]')
        mm.append(f"  {fam_id} --> {nid}")
        if any(n['hoja'] == 'PAPEL US$' for n in t['arbol_de_origen']):
            mm.append(f"  {nid} --> P")

    MMD.write_text("\n".join(mm), encoding="utf-8")

    print(json.dumps({
        "status": "ok",
        "total_traces": len(traces),
        "trees": len(traces),
        "families": sorted(set(FAMILY_MAP.values())),
        "papel_resueltas": len(resolved_papel_refs),
        "alerts": len(all_alerts),
        "top20": len(all_alerts_sorted[:20]),
        "files": [
            str(TRACE_JSON),
            str(ALERTS_JSON),
            str(DOC_TRACE),
            str(DOC_ALERTS),
            str(MMD),
        ],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
