import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parents[1]
EXCEL_PATH = BASE_DIR / "original" / "Copia de DIGITAL sistema 2025.xlsx"
OUT_DATA = BASE_DIR / "data" / "bajadas"
OUT_DOCS = BASE_DIR / "docs" / "bajadas"

OUT_DATA.mkdir(parents=True, exist_ok=True)
OUT_DOCS.mkdir(parents=True, exist_ok=True)

ERROR_TOKENS = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!"}


def normalize_formula(formula: str) -> str:
    f = formula
    if f.startswith("="):
        f = f[1:]

    def repl(match):
        col = match.group(1)
        row = int(match.group(2))
        return f"{col}{{r{row}}}"

    return re.sub(r"\$?([A-Z]{1,3})\$?(\d+)", repl, f)


def sheet_refs_and_external(formula: str):
    sheet_refs = set()
    external_refs = set()

    for m in re.finditer(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_\. ]*))!", formula):
        name = m.group(1) or m.group(2)
        if name:
            sheet_refs.add(name.strip())

    for m in re.finditer(r"\[([^\]]+)\]([^!\[]+)!", formula):
        external_refs.add(f"[{m.group(1)}]{m.group(2)}")

    return sorted(sheet_refs), sorted(external_refs)


def detect_used_range(ws):
    min_row = None
    max_row = None
    min_col = None
    max_col = None

    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                if min_row is None or c.row < min_row:
                    min_row = c.row
                if max_row is None or c.row > max_row:
                    max_row = c.row
                if min_col is None or c.column < min_col:
                    min_col = c.column
                if max_col is None or c.column > max_col:
                    max_col = c.column

    if min_row is None:
        return None

    return {
        "min_row": min_row,
        "max_row": max_row,
        "min_col": min_col,
        "max_col": max_col,
        "address": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
    }


def detect_blocks(cells):
    by_row = defaultdict(list)
    for c in cells:
        by_row[c["row"]].append(c)

    row_ranges = {}
    for r, arr in by_row.items():
        cols = sorted(x["col"] for x in arr)
        row_ranges[r] = (min(cols), max(cols), len(cols))

    blocks = []
    current = None

    for r in sorted(row_ranges):
        cmin, cmax, count = row_ranges[r]
        if current is None:
            current = {"start_row": r, "end_row": r, "min_col": cmin, "max_col": cmax, "rows": 1}
        else:
            same_shape = abs(cmin - current["min_col"]) <= 1 and abs(cmax - current["max_col"]) <= 1
            consecutive = r == current["end_row"] + 1
            if consecutive and same_shape:
                current["end_row"] = r
                current["rows"] += 1
                current["min_col"] = min(current["min_col"], cmin)
                current["max_col"] = max(current["max_col"], cmax)
            else:
                blocks.append(current)
                current = {"start_row": r, "end_row": r, "min_col": cmin, "max_col": cmax, "rows": 1}

    if current:
        blocks.append(current)

    filtered = []
    for i, b in enumerate(blocks, 1):
        if b["rows"] >= 2:
            filtered.append({
                "id": f"B{i}",
                "range": f"{get_column_letter(b['min_col'])}{b['start_row']}:{get_column_letter(b['max_col'])}{b['end_row']}",
                "rows": b["rows"],
                "cols": b["max_col"] - b["min_col"] + 1,
            })

    return filtered


def classify_inputs_outputs(constants, formulas):
    input_candidates = []
    output_candidates = []

    formula_cells = {f["cell"] for f in formulas}
    referenced_cells = set()
    for f in formulas:
        for m in re.finditer(r"\$?([A-Z]{1,3})\$?(\d+)", f["formula"]):
            referenced_cells.add(f"{m.group(1)}{m.group(2)}")

    for c in constants:
        if c["cell"] in referenced_cells:
            input_candidates.append(c)

    for f in formulas:
        if f["cell"] not in referenced_cells:
            output_candidates.append(f)

    return input_candidates, output_candidates


def main():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=False, read_only=False)

    if "Bajadas" not in wb.sheetnames:
        raise ValueError("No existe la hoja 'Bajadas' en el Excel")

    ws = wb["Bajadas"]

    used_range = detect_used_range(ws)

    constants = []
    formulas = []
    errors = []
    dependencies_sheets = set()
    dependencies_external = set()
    all_non_empty = []

    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            val = c.value
            cell_ref = c.coordinate
            all_non_empty.append({"cell": cell_ref, "row": c.row, "col": c.column})

            if isinstance(val, str) and val.startswith("="):
                norm = normalize_formula(val)
                srefs, erefs = sheet_refs_and_external(val)
                for s in srefs:
                    if s != "Bajadas":
                        dependencies_sheets.add(s)
                for e in erefs:
                    dependencies_external.add(e)

                formulas.append({
                    "cell": cell_ref,
                    "row": c.row,
                    "col": c.column,
                    "formula": val,
                    "normalized_formula": norm,
                    "references_sheets": srefs,
                    "references_external": erefs,
                })
            else:
                constants.append({
                    "cell": cell_ref,
                    "row": c.row,
                    "col": c.column,
                    "value": val,
                    "type": type(val).__name__,
                })
                if isinstance(val, str) and val.strip().upper() in ERROR_TOKENS:
                    errors.append({
                        "cell": cell_ref,
                        "value": val.strip(),
                        "source": "literal_or_cached",
                    })

    formula_counter = Counter(f["normalized_formula"] for f in formulas)
    repeated_formulas = []
    for nf, cnt in formula_counter.items():
        if cnt > 1:
            cells = [f["cell"] for f in formulas if f["normalized_formula"] == nf]
            repeated_formulas.append({
                "normalized_formula": nf,
                "count": cnt,
                "cells": cells,
                "range_hint": f"{cells[0]}:{cells[-1]}" if len(cells) > 1 else cells[0],
            })

    repeated_formulas.sort(key=lambda x: x["count"], reverse=True)

    blocks = detect_blocks(all_non_empty)
    input_candidates, output_candidates = classify_inputs_outputs(constants, formulas)

    inventory = {
        "workbook": str(EXCEL_PATH),
        "sheet": "Bajadas",
        "used_range": used_range,
        "stats": {
            "non_empty_cells": len(all_non_empty),
            "constant_cells": len(constants),
            "formula_cells": len(formulas),
            "error_cells": len(errors),
            "repeated_formula_groups": len(repeated_formulas),
        },
        "possible_internal_blocks": blocks,
        "possible_input_cells": [c["cell"] for c in input_candidates[:200]],
        "possible_output_cells": [c["cell"] for c in output_candidates[:200]],
        "editable_or_business_constants": [
            {"cell": c["cell"], "value": c["value"], "type": c["type"]}
            for c in constants[:300]
        ],
        "status": "pendiente de interpretación",
    }

    formulas_out = {
        "sheet": "Bajadas",
        "total_formulas": len(formulas),
        "formulas": formulas,
        "repeated_formulas": repeated_formulas,
        "status": "pendiente de interpretación",
    }

    errors_out = {
        "sheet": "Bajadas",
        "total_errors": len(errors),
        "errors": errors,
        "status": "pendiente de interpretación",
    }

    dependencies_out = {
        "sheet": "Bajadas",
        "referenced_sheets": sorted(dependencies_sheets),
        "external_references": sorted(dependencies_external),
        "status": "pendiente de interpretación",
    }

    (OUT_DATA / "bajadas_inventory.json").write_text(json.dumps(inventory, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DATA / "bajadas_formulas.json").write_text(json.dumps(formulas_out, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DATA / "bajadas_errors.json").write_text(json.dumps(errors_out, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DATA / "bajadas_dependencies.json").write_text(json.dumps(dependencies_out, ensure_ascii=False, indent=2), encoding="utf-8")

    diag_md = []
    diag_md.append("# Diagnóstico Inicial - Hoja Bajadas")
    diag_md.append("")
    diag_md.append("Estado general: pendiente de interpretación")
    diag_md.append("")
    diag_md.append("## Inventario")
    diag_md.append(f"- Rango usado: `{used_range['address'] if used_range else 'N/A'}`")
    diag_md.append(f"- Celdas no vacías: `{len(all_non_empty)}`")
    diag_md.append(f"- Celdas con valores fijos: `{len(constants)}`")
    diag_md.append(f"- Celdas con fórmulas: `{len(formulas)}`")
    diag_md.append(f"- Celdas con errores detectables: `{len(errors)}`")
    diag_md.append("")
    diag_md.append("## Dependencias")
    diag_md.append(f"- Hojas referenciadas desde Bajadas: `{', '.join(sorted(dependencies_sheets)) if dependencies_sheets else 'Ninguna detectada'}`")
    diag_md.append(f"- Referencias externas [archivo]Hoja: `{', '.join(sorted(dependencies_external)) if dependencies_external else 'Ninguna detectada'}`")
    diag_md.append("")
    diag_md.append("## Bloques internos (heurístico)")
    if blocks:
        for b in blocks:
            diag_md.append(f"- {b['id']}: `{b['range']}` ({b['rows']} filas x {b['cols']} columnas)")
    else:
        diag_md.append("- No se pudieron inferir bloques con el heurístico actual.")
    diag_md.append("")
    diag_md.append("## Entradas y salidas (heurístico)")
    diag_md.append(f"- Posibles entradas (constantes referenciadas por fórmulas): `{len(input_candidates)}`")
    diag_md.append(f"- Posibles salidas (fórmulas no referenciadas por otras fórmulas): `{len(output_candidates)}`")
    diag_md.append("")
    diag_md.append("## Riesgos / dudas")
    diag_md.append("- La clasificación entrada/salida es heurística y requiere validación funcional.")
    diag_md.append("- Errores de fórmula no siempre son visibles sin recálculo en Excel; aquí se detectan errores presentes como valor literal/caché.")
    diag_md.append("- Falta interpretación semántica de negocio: pendiente de interpretación.")

    (OUT_DOCS / "00_diagnostico_bajadas.md").write_text("\n".join(diag_md), encoding="utf-8")

    tree_md = []
    tree_md.append("# Árbol Funcional Inicial - Bajadas")
    tree_md.append("")
    tree_md.append("Estado: pendiente de interpretación")
    tree_md.append("")
    tree_md.append("## Vista jerárquica")
    tree_md.append("1. Hoja `Bajadas`")
    tree_md.append("2. Bloques internos detectados (heurístico)")
    for b in blocks[:20]:
        tree_md.append(f"3. {b['id']} -> `{b['range']}`")
    if not blocks:
        tree_md.append("3. Sin bloques claros detectados por heurística")
    tree_md.append("2. Capa de cálculo")
    tree_md.append(f"3. Fórmulas totales: `{len(formulas)}`")
    tree_md.append(f"3. Patrones repetidos: `{len(repeated_formulas)}`")
    tree_md.append("2. Dependencias")
    if dependencies_sheets:
        for s in sorted(dependencies_sheets):
            tree_md.append(f"3. Hoja referenciada: `{s}`")
    else:
        tree_md.append("3. Sin referencias a otras hojas")
    if dependencies_external:
        for e in sorted(dependencies_external):
            tree_md.append(f"3. Referencia externa: `{e}`")
    else:
        tree_md.append("3. Sin referencias externas detectadas")

    (OUT_DOCS / "01_arbol_bajadas.md").write_text("\n".join(tree_md), encoding="utf-8")

    mermaid = []
    mermaid.append("flowchart TD")
    mermaid.append('  A["Bajadas"]')
    mermaid.append('  B["Bloques internos"]')
    mermaid.append('  C["Capa de fórmulas"]')
    mermaid.append('  D["Errores detectados"]')
    mermaid.append('  E["Dependencias"]')
    mermaid.append("  A --> B")
    mermaid.append("  A --> C")
    mermaid.append("  A --> D")
    mermaid.append("  A --> E")

    for i, b in enumerate(blocks[:10], 1):
        mermaid.append(f'  B{i}["{b["id"]}: {b["range"]}"]')
        mermaid.append(f"  B --> B{i}")

    mermaid.append(f'  C1["Formulas: {len(formulas)}"]')
    mermaid.append(f'  C2["Patrones repetidos: {len(repeated_formulas)}"]')
    mermaid.append("  C --> C1")
    mermaid.append("  C --> C2")
    mermaid.append(f'  D1["Errores: {len(errors)}"]')
    mermaid.append("  D --> D1")

    if dependencies_sheets:
        for i, s in enumerate(sorted(dependencies_sheets), 1):
            node = f"ES{i}"
            mermaid.append(f'  {node}["Hoja: {s}"]')
            mermaid.append(f"  E --> {node}")
    else:
        mermaid.append('  E0["Sin hojas externas"]')
        mermaid.append("  E --> E0")

    if dependencies_external:
        for i, e in enumerate(sorted(dependencies_external), 1):
            node = f"EX{i}"
            safe = e.replace('"', "'")
            mermaid.append(f'  {node}["Ref externa: {safe}"]')
            mermaid.append(f"  E --> {node}")

    (OUT_DOCS / "arbol_bajadas.mmd").write_text("\n".join(mermaid), encoding="utf-8")

    print(json.dumps({
        "status": "ok",
        "sheet": "Bajadas",
        "used_range": used_range,
        "formula_cells": len(formulas),
        "error_cells": len(errors),
        "dependencies_sheets": sorted(dependencies_sheets),
        "dependencies_external": sorted(dependencies_external),
        "output_dir": str(BASE_DIR),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
