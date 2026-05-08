import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

BASE = Path(__file__).resolve().parents[1]
EXCEL = BASE / "original" / "Copia de DIGITAL sistema 2025.xlsx"
DATA = BASE / "data" / "bajadas"
DOCS = BASE / "docs" / "bajadas"


def parse_range(rng: str):
    a, b = rng.split(":")
    m1 = re.match(r"([A-Z]+)(\d+)", a)
    m2 = re.match(r"([A-Z]+)(\d+)", b)
    c1, r1 = column_index_from_string(m1.group(1)), int(m1.group(2))
    c2, r2 = column_index_from_string(m2.group(1)), int(m2.group(2))
    return r1, c1, r2, c2


def in_range(row, col, rng):
    r1, c1, r2, c2 = parse_range(rng)
    return r1 <= row <= r2 and c1 <= col <= c2


def coord_to_rc(coord):
    m = re.match(r"([A-Z]+)(\d+)", coord)
    return int(m.group(2)), column_index_from_string(m.group(1))


def extract_cell_refs(formula):
    return [f"{m.group(1)}{m.group(2)}" for m in re.finditer(r"\$?([A-Z]{1,3})\$?(\d+)", formula)]


def classify_type(title, range_addr):
    t = (title or "").lower()
    if "coef" in t or "param" in t:
        return "tabla de referencia"
    if "costo" in t or "costo" in t:
        return "cálculo intermedio"
    if "bajada" in t or "a3" in t or "xa3" in t or "xl" in t:
        return "cálculo intermedio"
    if "encuadern" in t:
        return "salida"
    return "pendiente de interpretación"


def probable_function(title, block_id):
    t = (title or "").strip()
    if not t:
        return "PENDIENTE_DE_INTERPRETACION"
    low = t.lower()
    if "encuadern" in low:
        return "Matriz de precios de encuadernación por tipo y tramo (probable)."
    if "coef" in low or "param" in low:
        return "Parámetros/coeficientes de ajuste para cálculos de bajadas y costos (probable)."
    if "a3" in low or "xa3" in low or "a4" in low or "xl" in low or "bajada" in low:
        return "Cálculo de costos/precios por rangos de cantidad, formato, gramaje y caras (probable)."
    return "PENDIENTE_DE_INTERPRETACION"


def detect_headers(ws, rng):
    r1, c1, r2, c2 = parse_range(rng)
    headers = []
    for r in range(r1, min(r1 + 3, r2) + 1):
        for c in range(c1, c2 + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                txt = v.strip()
                if not txt.startswith("="):
                    headers.append((r, c, txt))
    headers.sort(key=lambda x: (x[0], x[1]))
    return headers


def top_formula_patterns(formulas, topn=5):
    def norm(f):
        return re.sub(r"\$?([A-Z]{1,3})\$?(\d+)", lambda m: f"{m.group(1)}{{r}}", f)

    ctr = Counter(norm(f["formula"]) for f in formulas)
    out = []
    for pat, cnt in ctr.most_common(topn):
        sample = next((x["formula"] for x in formulas if norm(x["formula"]) == pat), None)
        out.append({"pattern": pat, "count": cnt, "sample_formula": sample})
    return out


def main():
    DATA.mkdir(parents=True, exist_ok=True)
    DOCS.mkdir(parents=True, exist_ok=True)

    prev_inventory = json.loads((DATA / "bajadas_inventory.json").read_text(encoding="utf-8"))
    prev_formulas = json.loads((DATA / "bajadas_formulas.json").read_text(encoding="utf-8"))
    prev_dependencies = json.loads((DATA / "bajadas_dependencies.json").read_text(encoding="utf-8"))

    wb = load_workbook(EXCEL, data_only=False, read_only=True)
    ws = wb["Bajadas"]

    block_defs = [
        {"id": "SB01", "range": "W1:AJ1", "title_hint": "Parámetros iniciales tienda / coeficientes"},
        {"id": "SB02", "range": "B2:X11", "title_hint": "A3+ (32x47)"},
        {"id": "SB03", "range": "B12:X21", "title_hint": "XA3 (33x48)"},
        {"id": "SB04", "range": "B22:X31", "title_hint": "XL 32x70"},
        {"id": "SB05", "range": "Z16:AS44", "title_hint": "Parámetros / costos / sobres / clic"},
        {"id": "SB06", "range": "B32:X44", "title_hint": "Bajada A4 color"},
        {"id": "SB07", "range": "B45:AO53", "title_hint": "A3 PLUS ByN y rango pack color"},
        {"id": "SB08", "range": "B54:AG62", "title_hint": "XA3 ByN y rango pack ByN"},
        {"id": "SB09", "range": "B63:AC71", "title_hint": "Bajada XL ByN"},
        {"id": "SB10", "range": "AB69:AJ71", "title_hint": "Tabla auxiliar A3 PLUS (detalle)"},
        {"id": "SB11", "range": "B72:AB79", "title_hint": "Bajada A4 ByN"},
        {"id": "SB12", "range": "B84:E92", "title_hint": "Encuadernación"},
    ]

    all_cells = []
    formula_cells = []
    for r in range(1, 93):
        for c in range(1, 49):
            v = ws.cell(r, c).value
            if v is None:
                continue
            cell = f"{get_column_letter(c)}{r}"
            rec = {"cell": cell, "row": r, "col": c, "value": v}
            all_cells.append(rec)
            if isinstance(v, str) and v.startswith("="):
                formula_cells.append(rec)

    block_index = {}
    for b in block_defs:
        r1, c1, r2, c2 = parse_range(b["range"])
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                block_index[(rr, cc)] = b["id"]

    formulas_map = {f["cell"]: f for f in prev_formulas["formulas"]}

    block_cards = []
    all_inputs = []
    all_outputs = []

    refs_to_papel = 0
    for b in block_defs:
        r1, c1, r2, c2 = parse_range(b["range"])
        block_cells = [x for x in all_cells if in_range(x["row"], x["col"], b["range"])]
        block_formulas = []
        for x in block_cells:
            if isinstance(x["value"], str) and x["value"].startswith("="):
                fobj = formulas_map.get(x["cell"], {"cell": x["cell"], "formula": x["value"], "references_sheets": [], "references_external": []})
                block_formulas.append(fobj)

        headers = detect_headers(ws, b["range"])
        detected_title = headers[0][2] if headers else b["title_hint"]

        internal_deps = set()
        deps_papel = set()
        produced_outputs = []

        referenced_within_block = set()
        for f in block_formulas:
            refs = extract_cell_refs(f["formula"])
            for ref in refs:
                rr, cc = coord_to_rc(ref)
                bid = block_index.get((rr, cc))
                if bid:
                    if bid != b["id"]:
                        internal_deps.add(bid)
                    else:
                        referenced_within_block.add(ref)
            if "PAPEL US$" in f.get("references_sheets", []):
                deps_papel.add(f["cell"])
                refs_to_papel += 1

        # Entradas probables: constantes numéricas o texto-rango en columna inicial
        input_candidates = []
        for c in block_cells:
            v = c["value"]
            if isinstance(v, str) and not v.startswith("="):
                low = v.lower()
                if any(k in low for k in [" a ", "g", "cm", "triplex", "sticker", "obra", "opp", "imán", "4/0", "4/4", "1/0", "1/1"]):
                    input_candidates.append(c["cell"])
            elif isinstance(v, (int, float)) and c["col"] in (2, 22, 23, 24, 25, 26, 27, 28):
                input_candidates.append(c["cell"])

        # Salidas probables: fórmulas no referenciadas por otras fórmulas del mismo bloque
        for f in block_formulas:
            if f["cell"] not in referenced_within_block:
                produced_outputs.append(f["cell"])

        block_type = classify_type(detected_title, b["range"])
        fn_prob = probable_function(detected_title, b["id"])

        # etiquetas de columnas/filas de negocio (heurístico textual)
        business_axis = set()
        for c in block_cells:
            v = c["value"]
            if isinstance(v, str) and not v.startswith("="):
                t = v.lower()
                mapping = {
                    "cantidad": ["cantidad", " a "],
                    "gramajes": ["g", "grs", "gram"],
                    "tamaños": ["x", "cm", "a3", "a4", "xa3", "xl"],
                    "caras_impresion": ["4/0", "4/4", "1/0", "1/1"],
                    "coste_papel": ["costo de papel", "costo de pliego", "papel"],
                    "coste_impresion": ["costo de clic", "clic"],
                    "precio_base": ["costo de bajada"],
                    "margen": ["coef", "ganacia", "ganancia"],
                    "desperdicio": ["desperd", "merma"],
                    "bajadas_por_pliego": ["x pliegos", "bajadas"],
                    "pliegos_necesarios": ["pliego"],
                    "precio_unitario": ["unit"],
                    "precio_final": ["precio", "final"],
                    "rangos_produccion": ["rango", " a "],
                }
                for k, vals in mapping.items():
                    if any(x in t for x in vals):
                        business_axis.add(k)

        all_inputs.extend(input_candidates)
        if block_type in ("salida", "cálculo intermedio"):
            all_outputs.extend(produced_outputs)

        card = {
            "id": b["id"],
            "range": b["range"],
            "title_detected": detected_title,
            "function_probable": fn_prob,
            "block_type": block_type,
            "non_empty_cells": len(block_cells),
            "formula_cells": len(block_formulas),
            "representative_formulas": top_formula_patterns(block_formulas, topn=5),
            "dependencies_internal_bajadas": sorted(internal_deps),
            "dependencies_external_papel_usd": sorted(deps_papel),
            "possible_outputs_generated": sorted(set(produced_outputs))[:50],
            "possible_inputs_detected": sorted(set(input_candidates))[:50],
            "detected_business_dimensions": sorted(business_axis),
            "pending_questions": [
                "PENDIENTE_DE_INTERPRETACION: Validar semántica exacta de columnas y unidades.",
                "PENDIENTE_DE_INTERPRETACION: Confirmar si las fórmulas finales representan precio final, unitario o costo intermedio.",
            ],
            "status": "PENDIENTE_DE_INTERPRETACION",
        }
        block_cards.append(card)

    # entradas/salidas globales probables por lectura semántica de títulos/celdas
    probable_inputs = sorted(set(all_inputs))
    probable_outputs = sorted(set(all_outputs))

    semantic = {
        "sheet": "Bajadas",
        "source_files": [
            "data/bajadas/bajadas_inventory.json",
            "data/bajadas/bajadas_formulas.json",
            "data/bajadas/bajadas_dependencies.json",
        ],
        "used_range": prev_inventory.get("used_range", {}),
        "declared_dependency": prev_dependencies.get("referenced_sheets", []),
        "blocks": block_cards,
        "global_business_dimensions_detected": sorted({d for b in block_cards for d in b["detected_business_dimensions"]}),
        "global_probable_inputs": probable_inputs[:300],
        "global_probable_outputs": probable_outputs[:300],
        "papel_usd_reference_cells_count": refs_to_papel,
        "status": "PENDIENTE_DE_INTERPRETACION",
        "notes": [
            "No se modificó el Excel original.",
            "No se analizaron otras hojas en profundidad; solo dependencia referenciada PAPEL US$.",
            "Clasificación funcional con heurística semántica de segunda pasada.",
        ],
    }

    (DATA / "bajadas_blocks_semantic.json").write_text(json.dumps(semantic, ensure_ascii=False, indent=2), encoding="utf-8")

    # 02 mapa funcional
    lines = []
    lines.append("# 02 - Mapa Funcional de Bajadas (Segunda Pasada)")
    lines.append("")
    lines.append("Estado: PENDIENTE_DE_INTERPRETACION")
    lines.append("")
    lines.append("## Alcance")
    lines.append("- Hoja analizada: `Bajadas`")
    lines.append("- Dependencia externa referenciada: `PAPEL US$` (sin análisis profundo)")
    lines.append("- Excel original: sin modificaciones")
    lines.append("")
    lines.append("## Bloques semánticos detectados")
    lines.append("| ID | Rango | Tipo | Título detectado | Función probable | Celdas con valor | Fórmulas |")
    lines.append("|---|---|---|---|---|---:|---:|")
    for b in block_cards:
        lines.append(f"| {b['id']} | {b['range']} | {b['block_type']} | {b['title_detected']} | {b['function_probable']} | {b['non_empty_cells']} | {b['formula_cells']} |")

    lines.append("")
    lines.append("## Dimensiones de negocio detectadas (heurístico)")
    for d in semantic["global_business_dimensions_detected"]:
        lines.append(f"- {d}")

    lines.append("")
    lines.append("## Dependencias")
    lines.append(f"- Hojas referenciadas: {', '.join(prev_dependencies.get('referenced_sheets', [])) or 'Ninguna'}")
    lines.append(f"- Celdas con referencia a PAPEL US$: `{refs_to_papel}`")
    lines.append("")
    lines.append("## Dudas")
    lines.append("- PENDIENTE_DE_INTERPRETACION: mapear inequívocamente precio base/final/unitario por columna en cada sub-bloque.")
    lines.append("- PENDIENTE_DE_INTERPRETACION: validar si los coeficientes son margen comercial, desperdicio o ajuste técnico según contexto de producto.")

    (DOCS / "02_mapa_funcional_bajadas.md").write_text("\n".join(lines), encoding="utf-8")

    # 03 entradas salidas
    eio = []
    eio.append("# 03 - Entradas y Salidas de Bajadas")
    eio.append("")
    eio.append("Estado: PENDIENTE_DE_INTERPRETACION")
    eio.append("")
    eio.append("## Entradas reales o probables")
    eio.append("- Rango/cantidad (ej: `2 a 25`, `26 a 50`, `501 a 1000`)")
    eio.append("- Gramajes y tipo de papel (Triplex, Obra, Sticker, OPP, Imán)")
    eio.append("- Tamaños/formato (A3, XA3, A4, XL, medidas cm)")
    eio.append("- Caras de impresión (`4/0`, `4/4`, `1/0`, `1/1`)")
    eio.append("- Coeficientes y parámetros (rangos pack, coeficientes de color/ByN, clic)")
    eio.append("- Dependencia de costos de papel referenciando `PAPEL US$`")
    eio.append("")
    eio.append("## Salidas reales o probables")
    eio.append("- Celdas de cálculo final por tramo dentro de cada bloque de bajada")
    eio.append("- Matrices de costo/precio en zonas derechas y tablas de cada formato")
    eio.append("- Tabla de encuadernación (`B84:E92`) como salida tarifaria probable")
    eio.append("")
    eio.append("## Listado técnico (muestra)")
    eio.append(f"- Entradas candidatas detectadas: `{len(probable_inputs)}`")
    eio.append(f"- Salidas candidatas detectadas: `{len(probable_outputs)}`")
    eio.append(f"- Ejemplo entradas: `{', '.join(probable_inputs[:20])}`")
    eio.append(f"- Ejemplo salidas: `{', '.join(probable_outputs[:20])}`")
    eio.append("")
    eio.append("## Dudas pendientes")
    eio.append("- PENDIENTE_DE_INTERPRETACION: confirmar qué columnas son precio base vs precio final vs precio unitario.")
    eio.append("- PENDIENTE_DE_INTERPRETACION: confirmar fórmula exacta de margen/desperdicio por familia de producto.")

    (DOCS / "03_entradas_salidas_bajadas.md").write_text("\n".join(eio), encoding="utf-8")

    # mermaid semántico
    mm = []
    mm.append("flowchart TD")
    mm.append('  A["Bajadas"]')
    mm.append('  E["Entradas detectadas"]')
    mm.append('  T["Tablas internas"]')
    mm.append('  C["Cálculos intermedios"]')
    mm.append('  D["Dependencia PAPEL US$"]')
    mm.append('  S["Salidas detectadas"]')
    mm.append("  A --> E")
    mm.append("  A --> T")
    mm.append("  A --> C")
    mm.append("  A --> D")
    mm.append("  A --> S")

    for b in block_cards:
        node = b["id"]
        label = f"{b['id']} {b['range']}\\n{b['block_type']}"
        mm.append(f'  {node}["{label}"]')
        if b["block_type"] == "tabla de referencia":
            mm.append(f"  T --> {node}")
        elif b["block_type"] == "cálculo intermedio":
            mm.append(f"  C --> {node}")
        elif b["block_type"] == "salida":
            mm.append(f"  S --> {node}")
        elif b["block_type"] == "entrada":
            mm.append(f"  E --> {node}")
        else:
            mm.append(f"  T --> {node}")

        if b["dependencies_external_papel_usd"]:
            mm.append(f"  {node} --> D")

    mm.append('  P["PENDIENTE_DE_INTERPRETACION"]')
    mm.append("  A --> P")

    (DOCS / "arbol_bajadas_semantico.mmd").write_text("\n".join(mm), encoding="utf-8")

    print(json.dumps({
        "status": "ok",
        "created": [
            "data/bajadas/bajadas_blocks_semantic.json",
            "docs/bajadas/02_mapa_funcional_bajadas.md",
            "docs/bajadas/03_entradas_salidas_bajadas.md",
            "docs/bajadas/arbol_bajadas_semantico.mmd",
        ],
        "blocks": len(block_cards),
        "papel_refs": refs_to_papel,
        "inputs": len(probable_inputs),
        "outputs": len(probable_outputs),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
