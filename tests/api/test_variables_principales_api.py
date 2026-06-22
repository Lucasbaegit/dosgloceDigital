import json
import sys
import threading
import unittest
from io import BytesIO
from pathlib import Path
from urllib import error, request

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "src"))

from api.main import create_server


class TestVariablesPrincipalesApi(unittest.TestCase):
    REQUIRED_EXCEL_SHEETS = [
        "00_RESUMEN",
        "01_VARIABLES_MADRE",
        "02_RANGOS",
        "03_BAJADAS_FULLCOLOR",
        "04_BAJADAS_BYN",
        "05_AUTOADHESIVAS",
        "06_TROQUELADO_DIGITAL",
        "07_STICKERS_CIRCULARES",
        "08_TARJETAS_TROQUELADAS",
        "09_TARJETAS_9X5",
        "10_TARJETAS_POSTALES",
        "11_FOLLETOS",
        "12_CARPETAS",
        "13_SOBRES",
        "14_STICKERS_CORTE_RECTO",
        "15_IMANES_CORTE_RECTO",
        "16_PLANCHA_IMAN",
        "17_AGENDAS_CUADERNOS",
        "18_BLOQUEADOS",
        "19_TRAZABILIDAD",
        "21_TRAZABILIDAD_PRECIOS",
    ]

    @classmethod
    def setUpClass(cls):
        cls.paths = [
            ROOT / "data" / "bajadas_v2" / "bajadas_v2_config_final.json",
            ROOT / "data" / "stickers_circulares" / "formula_editable_config.json",
            ROOT / "data" / "bajadas_autoadhesivas" / "autoadhesivas_v1_config.json",
            ROOT / "data" / "variables_principales" / "variables_madre.json",
        ]
        cls.originals = {path: path.read_text(encoding="utf-8") for path in cls.paths}
        cls.backup_dir = ROOT / "data" / "backups" / "variables_principales"
        cls.backups_before = set(cls.backup_dir.glob("*.json")) if cls.backup_dir.exists() else set()
        cls.server = create_server(host="127.0.0.1", port=0, project_root=ROOT)
        cls.port = cls.server.server_address[1]
        cls.thread = threading.Thread(target=cls.server.serve_forever, daemon=True)
        cls.thread.start()

    @classmethod
    def tearDownClass(cls):
        cls.server.shutdown()
        cls.server.server_close()
        cls.thread.join(timeout=2)
        for path, content in cls.originals.items():
            path.write_text(content, encoding="utf-8")
        if cls.backup_dir.exists():
            for path in set(cls.backup_dir.glob("*.json")) - cls.backups_before:
                path.unlink()

    def call(self, method, path, payload=None):
        data = None if payload is None else json.dumps(payload).encode("utf-8")
        req = request.Request(
            f"http://127.0.0.1:{self.port}{path}",
            data=data,
            headers={"Content-Type": "application/json"},
            method=method,
        )
        try:
            with request.urlopen(req, timeout=5) as resp:
                return resp.status, json.loads(resp.read().decode("utf-8"))
        except error.HTTPError as exc:
            try:
                return exc.code, json.loads(exc.read().decode("utf-8"))
            finally:
                exc.close()

    def call_raw(self, path):
        with request.urlopen(f"http://127.0.0.1:{self.port}{path}", timeout=10) as resp:
            return resp.status, dict(resp.headers.items()), resp.read()

    def test_get_devuelve_solo_catalogo_controlado(self):
        status, body = self.call("GET", "/variables-principales")
        self.assertEqual(status, 200)
        keys = {
            item["key"]
            for group in ["tipo_cambio", "clicks", "papeles", "multiplicadores", "adicionales"]
            for item in body[group]
        }
        self.assertIn("tipo_cambio_usd", keys)
        self.assertIn("click_color", keys)
        self.assertIn("click_bn_excel", keys)
        self.assertIn("ilustracion_150g_65x95_usd", keys)
        self.assertIn("adicional_tinta_blanca_base_1_copia", keys)
        self.assertIn("laca_uv_factor_stickers_circulares", keys)
        self.assertIn("corte_circular_factor_stickers_circulares", keys)
        self.assertIn("multiplicador_comercial_stickers_circulares", keys)
        self.assertIn("coeficiente_tamano_stickers_circulares_10cm", keys)
        self.assertIn("coeficiente_cantidad_stickers_circulares_1000", keys)
        self.assertNotIn("factor_ajuste_pdf", keys)
        editable_items = [
            item
            for group in ["tipo_cambio", "clicks", "papeles", "multiplicadores", "adicionales"]
            for item in body[group]
        ]
        self.assertGreater(len(editable_items), 5)
        self.assertTrue(all(item["editable"] for item in editable_items))
        self.assertTrue(all(item["tipo"] == "variable_madre" for item in editable_items))
        self.assertTrue(all("impacta_hoy" in item for item in editable_items))
        self.assertTrue(all(item.get("value") is not None for item in editable_items))
        self.assertTrue(all(item.get("confiabilidad") in {"alta", "media"} for item in editable_items))
        self.assertTrue(any(item["key"] == "click_color" and item["impacta_hoy"] for item in editable_items))
        self.assertTrue(any(item["key"] == "laca_uv_factor_stickers_circulares" and item["impacta_hoy"] for item in editable_items))
        self.assertTrue(any(item["key"] == "click_bn_excel" and not item["impacta_hoy"] for item in editable_items))
        obra_90g = next(item for item in editable_items if item["key"] == "obra_90g")
        self.assertEqual(obra_90g["productos_afectados"], ["Stickers Circulares"])
        self.assertIn("No está conectado a Bajadas Autoadhesivas", obra_90g["description"])
        self.assertIn("Papeles Bajadas", body["papeles_detectados"])
        self.assertIn("Papeles Stickers Circulares", body["papeles_detectados"])
        self.assertTrue(any(item["key"] == "obra_90g" and item["tipo"] == "variable_madre" and item["editable"] for item in body["papeles_detectados"]["Papeles Stickers Circulares"]))
        self.assertTrue(any(item["key"] == "obra_90g" and item["tipo"] == "detectado_sin_costo_base" and not item["editable"] for item in body["papeles_detectados"]["Papeles Bajadas"]))
        self.assertFalse(any(item["key"] == "obra_90g" and item["editable"] for item in body["papeles_detectados"]["Papeles Autoadhesivas"]))
        self.assertTrue(any(item["key"] == "ilustracion_150g" and item["tipo"] == "detectado_sin_costo_base" and not item["editable"] for item in body["papeles_detectados"]["Papeles Bajadas"]))
        self.assertTrue(body["variables_madre_impactan_hoy"])
        self.assertTrue(body["variables_madre_preparadas"])
        self.assertTrue(all(not item["editable"] for item in body["valores_derivados"]))
        self.assertTrue(all(item["tipo"] == "tabla_fija_pdf" and not item["editable"] for item in body["tablas_fijas_pdf"]))

    def test_rangos_son_visibles_y_no_editables(self):
        status, body = self.call("GET", "/variables-principales/rangos")
        self.assertEqual(status, 200)
        self.assertGreaterEqual(len(body["rangos"]), 9)
        self.assertTrue(all(not item["editable"] for item in body["rangos"]))
        self.assertTrue(all(item["tipo"] == "rango_fijo" for item in body["rangos"]))
        self.assertTrue(all("motivo" in item for item in body["rangos"]))

    def test_export_json_y_pdf(self):
        status, body = self.call("GET", "/export/precios/json")
        self.assertEqual(status, 200)
        titles = {item["titulo"] for item in body["tablas"]}
        for expected in ["Tarjetas Personales", "Tarjetas Postales", "Folletos", "Carpetas", "Sobres", "Stickers Corte Recto", "Imanes Corte Recto", "Plancha de Imán Impreso", "Agendas / Cuadernos"]:
            self.assertIn(expected, titles)
        self.assertEqual(next(item for item in body["tablas"] if item["titulo"] == "Membretes")["estado"], "bloqueado")
        self.assertNotIn("DTF UV", titles)

        status, headers, pdf = self.call_raw("/export/precios/pdf")
        self.assertEqual(status, 200)
        self.assertEqual(headers["Content-Type"], "application/pdf")
        self.assertTrue(headers["Content-Disposition"].endswith('.pdf"'))
        self.assertGreater(len(pdf), 1000)
        self.assertTrue(pdf.startswith(b"%PDF-"))

    def test_export_excel_maestro(self):
        status, headers, payload = self.call_raw("/export/precios/excel")
        self.assertEqual(status, 200)
        self.assertEqual(
            headers["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("cotizador_digital_maestro_", headers["Content-Disposition"])
        self.assertTrue(headers["Content-Disposition"].endswith('.xlsx"'))
        self.assertGreater(len(payload), 1000)

        workbook = load_workbook(BytesIO(payload), data_only=True)
        self.assertEqual(workbook.sheetnames, self.REQUIRED_EXCEL_SHEETS)

        variables = workbook["01_VARIABLES_MADRE"]
        headers_row = [cell.value for cell in variables[1]]
        self.assertEqual(headers_row, [
            "key",
            "label",
            "value",
            "unit",
            "editable",
            "editable_en_sistema",
            "editable_en_excel_maestro",
            "tipo",
            "impacta_hoy",
            "estado_operativo",
            "productos_afectados",
            "source_file",
            "source_path",
            "source_sheet",
            "source_cell",
            "description",
        ])
        variable_keys = [row[0] for row in variables.iter_rows(min_row=2, values_only=True) if row[0]]
        self.assertIn("adicional_tinta_blanca_base_1_copia", variable_keys)
        self.assertFalse(any("precio_sin_iva" in str(key) for key in variable_keys))
        self.assertFalse(any("precio_pdf" in str(key) for key in variable_keys))
        self.assertFalse(any("factor_ajuste_pdf" in str(key) for key in variable_keys))
        variable_rows = {row[0]: dict(zip(headers_row, row)) for row in variables.iter_rows(min_row=2, values_only=True) if row[0]}
        self.assertTrue(variable_rows["click_color"]["editable_en_sistema"])
        self.assertEqual(variable_rows["click_color"]["estado_operativo"], "operativa")
        self.assertFalse(variable_rows["ilustracion_150g_65x95_usd"]["editable_en_sistema"])
        self.assertTrue(variable_rows["ilustracion_150g_65x95_usd"]["editable_en_excel_maestro"])
        self.assertEqual(variable_rows["ilustracion_150g_65x95_usd"]["estado_operativo"], "preparada_no_conectada")

        blocked_values = {
            str(value)
            for row in workbook["18_BLOQUEADOS"].iter_rows(values_only=True)
            for value in row
            if value is not None
        }
        self.assertIn("Membretes", blocked_values)
        self.assertIn("OPP Stickers Circulares", blocked_values)
        self.assertIn("DTF UV", blocked_values)
        self.assertTrue(any(value.startswith("PegaMan") for value in blocked_values))
        blocked_headers = [cell.value for cell in workbook["18_BLOQUEADOS"][1]]
        self.assertEqual(blocked_headers, ["producto", "estado", "motivo", "fuente", "accion_necesaria", "precio_inventado"])

        trace_values = {
            str(value)
            for row in workbook["19_TRAZABILIDAD"].iter_rows(values_only=True)
            for value in row
            if value is not None
        }
        self.assertIn("/bajadas-v2/cotizar", trace_values)
        self.assertIn("/stickers-circulares/cotizar", trace_values)
        trace_headers = [cell.value for cell in workbook["19_TRAZABILIDAD"][1]]
        self.assertEqual(trace_headers, ["producto", "endpoint", "modo_precio", "fuente_pdf", "fuente_excel", "archivo_data", "motor_backend", "estado", "notas"])

        price_trace = workbook["21_TRAZABILIDAD_PRECIOS"]
        price_trace_headers = [cell.value for cell in price_trace[1]]
        self.assertIn("producto", price_trace_headers)
        self.assertIn("componente", price_trace_headers)
        self.assertIn("tipo_componente", price_trace_headers)
        self.assertIn("editable_en_sistema", price_trace_headers)
        self.assertIn("impacta_hoy", price_trace_headers)
        price_rows = [dict(zip(price_trace_headers, row)) for row in price_trace.iter_rows(min_row=2, values_only=True) if row[0]]
        self.assertTrue(any(row["producto"] == "Tarjetas 9x5" and row["precio_final"] == 5139 for row in price_rows))
        self.assertTrue(any(row["producto"] == "Stickers Circulares" and row["precio_final"] == 85980 for row in price_rows))
        self.assertTrue(any(row["producto"] == "Bajadas Autoadhesivas" and row["componente"] == "tinta_blanca" and row["valor_base"] == 603 for row in price_rows))
        self.assertTrue(any(row["producto"] == "Membretes" and row["modo_precio"] == "bloqueado" for row in price_rows))
        self.assertTrue(any(row["tipo_componente"] == "tabla_pdf" for row in price_rows))
        self.assertTrue(any(row["tipo_componente"] == "variable_madre" for row in price_rows))
        self.assertTrue(any(row["tipo_componente"] == "adicional_variable" for row in price_rows))
        self.assertFalse(any(row["tipo_componente"] == "tabla_pdf" and row["editable_en_sistema"] for row in price_rows))
        self.assertFalse(any(row["estado_operativo"] == "tabla_pdf_fija" and row["impacta_hoy"] for row in price_rows))

    def test_put_rechaza_key_y_valores_invalidos(self):
        status, _ = self.call("PUT", "/variables-principales", {"updates": [{"key": "factor_ajuste_pdf", "value": 2}]})
        self.assertEqual(status, 400)
        status, _ = self.call("PUT", "/variables-principales", {"updates": [{"key": "click_color", "value": -1}]})
        self.assertEqual(status, 400)
        status, _ = self.call("PUT", "/variables-principales", {"updates": [{"key": "click_color", "value": "texto"}]})
        self.assertEqual(status, 400)

    def test_put_permite_variable_madre_preparada(self):
        status, before = self.call("GET", "/variables-principales")
        self.assertEqual(status, 200)
        prepared = next(item for item in before["papeles"] if item["key"] == "ilustracion_150g_65x95_usd")
        status, body = self.call(
            "PUT",
            "/variables-principales",
            {"updates": [{"key": prepared["key"], "value": prepared["value"] + 0.001}]},
        )
        self.assertEqual(status, 200)
        self.assertEqual(len(body["backups_generados"]), 1)
        self.assertTrue(any(path.endswith("variables_madre.json") for path in body["backups_generados"]))

    def test_auditoria_incluye_excel_config_y_descartadas(self):
        status, body = self.call("GET", "/variables-principales/audit")
        self.assertEqual(status, 200)
        self.assertGreaterEqual(len(body["variables_madre_encontradas_config"]), 5)
        self.assertGreater(len(body["variables_madre_encontradas_excel"]), 0)
        self.assertTrue(any(item["key"] == "opp_clear_usd" for item in body["variables_descartadas_por_falta_de_valor"]))
        self.assertTrue(body["variables_preparadas_sin_impacto_actual"])

    def test_put_modifica_variables_y_crea_backup(self):
        status, before = self.call("GET", "/variables-principales")
        self.assertEqual(status, 200)
        values = {item["key"]: item["value"] for group in ["tipo_cambio", "clicks"] for item in before[group]}
        status, body = self.call(
            "PUT",
            "/variables-principales",
            {"updates": [
                {"key": "tipo_cambio_usd", "value": values["tipo_cambio_usd"] + 1},
                {"key": "click_color", "value": values["click_color"] + 0.1},
            ]},
        )
        self.assertEqual(status, 200)
        self.assertEqual(len(body["backups_generados"]), 2)
        self.assertTrue(all((ROOT / path).exists() for path in body["backups_generados"]))

    def test_tinta_blanca_actualizada_impacta_cotizacion(self):
        status, variables = self.call("GET", "/variables-principales")
        self.assertEqual(status, 200)
        original = variables["adicionales"][0]["value"]
        new_value = original + 10
        status, _ = self.call(
            "PUT",
            "/variables-principales",
            {"updates": [{"key": "adicional_tinta_blanca_base_1_copia", "value": new_value}]},
        )
        self.assertEqual(status, 200)
        payload = {
            "categoria": "Bajadas Autoadhesivas",
            "modo_color": "fullcolor",
            "formato": "A3+",
            "tipo_papel": "especial",
            "material": "OPP blanco",
            "gramaje": "N/A",
            "cantidad_unidades": 30,
            "cantidad_rango": "26 a 50",
            "caras": "4/0",
            "urgencia": "normal",
            "tipo_producto": "autoadhesiva",
            "columna_precio": "especial",
            "adicional_tinta_blanca": True,
        }
        status, quote = self.call("POST", "/bajadas-v2/cotizar", payload)
        self.assertEqual(status, 200)
        self.assertAlmostEqual(quote["total_adicional_sin_iva"], new_value * 30, places=6)
        self.assertEqual(
            quote["trazabilidad"]["variables_principales_usadas"][0]["key"],
            "adicional_tinta_blanca_base_1_copia",
        )


if __name__ == "__main__":
    unittest.main()
