"""Preview-only importer for master price Excel workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pricing_variables import PrincipalVariablesService


REQUIRED_COLUMNS = {"key", "value", "unit", "editable_en_sistema", "impacta_hoy", "estado_operativo"}
IMPORTABLE_KEYS = {
    "tipo_cambio_usd",
    "click_color",
    "obra_90g",
    "multiplicador_general",
    "adicional_tinta_blanca_base_1_copia",
}
FORBIDDEN_KEY_TOKENS = (
    "precio_final",
    "precio_pdf",
    "precio_objetivo_pdf",
    "factor_ajuste_pdf",
    "total_final",
    "matriz_pdf",
)


@dataclass(frozen=True)
class ExcelMasterPreviewError(Exception):
    error: str
    detail: str


class ExcelMaestroImporter:
    def __init__(self, project_root: Path):
        self.project_root = project_root
        self.variables_service = PrincipalVariablesService(project_root)

    def preview(self, filename: str, content: bytes) -> dict[str, Any]:
        if not filename.lower().endswith(".xlsx"):
            raise ExcelMasterPreviewError("archivo_invalido", "El archivo debe tener extension .xlsx")
        try:
            workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        except Exception as exc:
            raise ExcelMasterPreviewError("excel_no_parseable", f"No se pudo abrir el Excel: {exc}") from exc

        if "01_VARIABLES_MADRE" not in workbook.sheetnames:
            raise ExcelMasterPreviewError("hoja_requerida_faltante", "Falta la hoja 01_VARIABLES_MADRE")
        sheet = workbook["01_VARIABLES_MADRE"]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ExcelMasterPreviewError("hoja_vacia", "La hoja 01_VARIABLES_MADRE no contiene filas de variables")

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        missing = sorted(REQUIRED_COLUMNS - set(headers))
        if missing:
            raise ExcelMasterPreviewError("columnas_requeridas_faltantes", f"Faltan columnas requeridas: {', '.join(missing)}")
        if len(rows) < 2:
            raise ExcelMasterPreviewError("hoja_vacia", "La hoja 01_VARIABLES_MADRE no contiene filas de variables")

        header_index = {header: index for index, header in enumerate(headers) if header}
        current = self._current_variables_by_key()
        cambios: list[dict[str, Any]] = []
        bloqueados: list[dict[str, Any]] = []
        errores: list[dict[str, Any]] = []
        advertencias: list[dict[str, Any]] = []
        variables_leidas = 0

        for row_number, row in enumerate(rows[1:], start=2):
            key = self._cell(row, header_index, "key")
            if key is None or str(key).strip() == "":
                if any(value is not None for value in row):
                    advertencias.append({"fila": row_number, "motivo": "key_vacia", "detail": "Fila ignorada por key vacia"})
                continue
            key = str(key).strip()
            variables_leidas += 1
            excel_value = self._cell(row, header_index, "value")
            label = self._cell(row, header_index, "label") or key
            unit = self._cell(row, header_index, "unit")
            excel_state = {
                "editable_en_sistema": self._as_bool(self._cell(row, header_index, "editable_en_sistema")),
                "impacta_hoy": self._as_bool(self._cell(row, header_index, "impacta_hoy")),
                "estado_operativo": str(self._cell(row, header_index, "estado_operativo") or ""),
            }

            if key == "rango" or any(token in key for token in FORBIDDEN_KEY_TOKENS):
                bloqueados.append(self._blocked_payload(key, label, None, excel_value, unit, excel_state, "No se importan precios finales, matrices, rangos ni derivados."))
                continue

            system = current.get(key)
            if system is None:
                advertencias.append({"fila": row_number, "key": key, "motivo": "key_desconocida", "detail": "La variable no existe en el sistema actual."})
                continue

            actual_value = system.get("value")
            if excel_value is None or str(excel_value).strip() == "":
                bloqueados.append(self._blocked_payload(key, label, actual_value, excel_value, unit, excel_state, "Valor vacio. No se importa en preview."))
                continue
            if isinstance(excel_value, bool) or not isinstance(excel_value, (int, float)):
                bloqueados.append(self._blocked_payload(key, label, actual_value, excel_value, unit, excel_state, "Valor no numerico."))
                continue
            if excel_value < 0:
                bloqueados.append(self._blocked_payload(key, label, actual_value, excel_value, unit, excel_state, "Valor negativo no permitido."))
                continue

            importable = self._system_importable(system)
            changed = not self._same_number(actual_value, excel_value)
            if not changed:
                continue
            if not importable:
                bloqueados.append(self._blocked_payload(
                    key,
                    label,
                    actual_value,
                    excel_value,
                    unit,
                    excel_state,
                    "Variable preparada sin impacto actual. No se importa en esta etapa.",
                    system,
                ))
                continue
            cambios.append({
                "key": key,
                "label": system.get("label") or label,
                "valor_actual": actual_value,
                "valor_excel": excel_value,
                "diferencia": excel_value - actual_value if isinstance(actual_value, (int, float)) else None,
                "unit": system.get("unit") or unit,
                "editable_en_sistema": True,
                "impacta_hoy": True,
                "estado_operativo": "operativa",
                "importable": True,
                "productos_afectados": system.get("productos_afectados", []),
                "motivo": "Variable madre operativa",
            })

        return {
            "ok": True,
            "archivo": filename,
            "hoja_leida": "01_VARIABLES_MADRE",
            "resumen": {
                "variables_leidas": variables_leidas,
                "cambios_detectados": len(cambios) + len(bloqueados),
                "cambios_importables": len(cambios),
                "cambios_bloqueados": len(bloqueados),
                "errores": len(errores),
                "advertencias": len(advertencias),
            },
            "cambios": cambios,
            "bloqueados": bloqueados,
            "errores": errores,
            "advertencias": advertencias,
        }

    def _current_variables_by_key(self) -> dict[str, dict[str, Any]]:
        grouped = self.variables_service.get_grouped()
        result: dict[str, dict[str, Any]] = {}
        for group in PrincipalVariablesService.GROUP_ORDER:
            for item in grouped.get(group, []):
                copied = dict(item)
                copied["editable_en_sistema"] = self._system_importable(copied)
                copied["estado_operativo"] = "operativa" if copied["editable_en_sistema"] else "preparada_no_conectada"
                result[copied["key"]] = copied
        return result

    def _system_importable(self, item: dict[str, Any]) -> bool:
        key = str(item.get("key") or "")
        return (
            key in IMPORTABLE_KEYS
            or key in {
                "laca_uv_factor_stickers_circulares",
                "corte_circular_factor_stickers_circulares",
                "multiplicador_comercial_stickers_circulares",
            }
            or key.startswith("coeficiente_tamano_stickers_circulares_")
            or key.startswith("coeficiente_cantidad_stickers_circulares_")
            or key in {
                "factor_laca_uv_stickers_corte_recto",
                "corte_recto_factor_stickers_corte_recto",
                "multiplicador_comercial_stickers_corte_recto",
                "factor_laca_uv_imanes_corte_recto",
                "corte_recto_factor_imanes_corte_recto",
                "multiplicador_comercial_imanes_corte_recto",
            }
            or key.startswith("coeficiente_formato_stickers_corte_recto_")
            or key.startswith("coeficiente_cantidad_stickers_corte_recto_")
            or key.startswith("coeficiente_formato_imanes_corte_recto_")
            or key.startswith("coeficiente_cantidad_imanes_corte_recto_")
            or key in {
                "factor_laca_uv_bajadas",
                "factor_troquelado_digital_bajadas",
                "factor_tinta_blanca_autoadhesivas",
                "multiplicador_comercial_bajadas",
                "factor_gramaje_tarjetas_9x5_350g",
                "factor_laca_uv_tarjetas_9x5",
                "factor_laminado_brillo_tarjetas_9x5",
                "factor_laminado_mate_tarjetas_9x5",
                "multiplicador_comercial_tarjetas_9x5",
                "factor_gramaje_tarjetas_postales_350g",
                "factor_laca_uv_tarjetas_postales",
                "factor_laminado_brillo_tarjetas_postales",
                "factor_laminado_mate_tarjetas_postales",
                "multiplicador_comercial_tarjetas_postales",
                "multiplicador_comercial_folletos",
            }
            or key.startswith("coeficiente_formato_bajadas_")
            or key.startswith("coeficiente_rango_bajadas_")
            or key.startswith("coeficiente_cantidad_tarjetas_9x5_")
            or key.startswith("coeficiente_impresion_tarjetas_9x5_")
            or key.startswith("coeficiente_cantidad_tarjetas_postales_")
            or key.startswith("coeficiente_impresion_tarjetas_postales_")
            or key.startswith("factor_papel_folletos_")
            or key.startswith("factor_formato_folletos_")
            or key.startswith("factor_color_folletos_")
            or key.startswith("factor_impresion_folletos_")
            or key.startswith("coeficiente_cantidad_folletos_")
            or key in {
                "factor_solapa_carpetas",
                "factor_laca_uv_carpetas",
                "factor_laminado_carpetas",
                "multiplicador_comercial_carpetas",
                "multiplicador_comercial_sobres",
                "base_iman_plancha",
                "papel_300g_ilustracion_plancha_iman",
                "multiplicador_comercial_plancha_iman",
                "base_agenda_2026",
                "factor_tapa_agendas",
                "factor_anillado_agendas",
                "multiplicador_comercial_agendas",
            }
            or key.startswith("coeficiente_terminacion_carpetas_")
            or key.startswith("coeficiente_impresion_carpetas_")
            or key.startswith("coeficiente_cantidad_carpetas_")
            or key.startswith("coeficiente_tipo_sobre_")
            or key.startswith("coeficiente_cantidad_sobres_")
            or key.startswith("coeficiente_variante_plancha_iman_")
            or key.startswith("coeficiente_cantidad_plancha_iman_")
            or key.startswith("coeficiente_producto_agendas_")
            or key.startswith("coeficiente_formato_agendas_")
            or key.startswith("coeficiente_paginas_agendas_")
        ) and (bool(item.get("impacta_hoy")) or key == "tipo_cambio_usd")

    def _blocked_payload(
        self,
        key: str,
        label: Any,
        actual_value: Any,
        excel_value: Any,
        unit: Any,
        excel_state: dict[str, Any],
        motivo: str,
        system: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        state_source = system or excel_state
        return {
            "key": key,
            "label": (system or {}).get("label") or label,
            "valor_actual": actual_value,
            "valor_excel": excel_value,
            "unit": (system or {}).get("unit") or unit,
            "editable_en_sistema": bool((system or {}).get("editable_en_sistema", False)),
            "impacta_hoy": bool((system or {}).get("impacta_hoy", excel_state.get("impacta_hoy", False))),
            "estado_operativo": (system or {}).get("estado_operativo") or state_source.get("estado_operativo") or "bloqueada",
            "importable": False,
            "motivo": motivo,
        }

    def _cell(self, row: tuple[Any, ...], index: dict[str, int], key: str) -> Any:
        position = index.get(key)
        if position is None or position >= len(row):
            return None
        return row[position]

    def _as_bool(self, value: Any) -> bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        return str(value).strip().lower() in {"true", "1", "si", "sí", "yes"}

    def _same_number(self, left: Any, right: Any) -> bool:
        if not isinstance(left, (int, float)) or isinstance(left, bool):
            return False
        return abs(float(left) - float(right)) <= 0.000001
