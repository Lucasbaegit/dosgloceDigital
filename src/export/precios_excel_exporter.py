"""Build the master editable-audit workbook for published price tables."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
import json
from pathlib import Path
import subprocess
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pricing_variables import PrincipalVariablesService, merge_global_base_costs


class PricesExcelExporter:
    MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    REQUIRED_SHEETS = [
        "00_RESUMEN",
        "01_VARIABLES_MADRE",
        "02_RANGOS",
        "03_BAJADAS_FULLCOLOR",
        "04_BAJADAS_BYN",
        "05_AUTOADHESIVAS",
        "06_TROQUELADO_DIGITAL",
        "07_STICKERS_CIRCULARES",
        "08_TARJETAS_TROQUELADAS",
        "09_TARJETAS_9X5",
        "10_TARJETAS_POSTALES",
        "11_FOLLETOS",
        "12_CARPETAS",
        "13_SOBRES",
        "14_STICKERS_CORTE_RECTO",
        "15_IMANES_CORTE_RECTO",
        "16_PLANCHA_IMAN",
        "17_AGENDAS_CUADERNOS",
        "18_BLOQUEADOS",
        "19_TRAZABILIDAD",
        "21_TRAZABILIDAD_PRECIOS",
    ]

    SHEET_BY_TITLE = {
        "Bajadas Autoadhesivas": "05_AUTOADHESIVAS",
        "Stickers Circulares": "07_STICKERS_CIRCULARES",
        "Troquelado Digital": "06_TROQUELADO_DIGITAL",
        "Tarjetas Troqueladas Circulares": "08_TARJETAS_TROQUELADAS",
        "Tarjetas Personales": "09_TARJETAS_9X5",
        "Tarjetas Postales": "10_TARJETAS_POSTALES",
        "Folletos": "11_FOLLETOS",
        "Carpetas": "12_CARPETAS",
        "Sobres": "13_SOBRES",
        "Stickers Corte Recto": "14_STICKERS_CORTE_RECTO",
        "Imanes Corte Recto": "15_IMANES_CORTE_RECTO",
        "Plancha de Imán Impreso": "16_PLANCHA_IMAN",
        "Plancha de Im?n Impreso": "16_PLANCHA_IMAN",
        "Agendas / Cuadernos": "17_AGENDAS_CUADERNOS",
    }

    BLOCKED_EXTRA = [
        {
            "producto": "OPP Stickers Circulares",
            "estado": "bloqueado",
            "motivo": "Material OPP pendiente por falta de matriz PDF/Excel confiable.",
            "accion_necesaria": "Aportar tabla o regla confiable para Stickers Circulares OPP.",
            "fuente": "data/stickers_circulares",
        },
        {
            "producto": "DTF UV",
            "estado": "excluido",
            "motivo": "Producto explícitamente excluido del alcance actual.",
            "accion_necesaria": "Definir alcance e implementación futura aprobada.",
            "fuente": "PDF/Excel",
        },
        {
            "producto": "DTF Textil",
            "estado": "excluido",
            "motivo": "Producto explícitamente excluido del alcance actual.",
            "accion_necesaria": "Definir alcance e implementación futura aprobada.",
            "fuente": "PDF/Excel",
        },
        {
            "producto": "PegaManía",
            "estado": "excluido",
            "motivo": "Producto explícitamente excluido del alcance actual.",
            "accion_necesaria": "Definir alcance e implementación futura aprobada.",
            "fuente": "PDF hojas 9 y 10",
        },
    ]

    LACA_UV_BAJADAS = {
        "1": 136,
        "2 a 25": 126,
        "26 a 50": 116,
        "51 a 100": 106,
        "101 a 300": 99,
        "301 a 500": 92,
        "501 a 1000": 85,
    }

    PRICE_TRACE_HEADERS = [
        "producto",
        "familia",
        "variante",
        "material",
        "gramaje",
        "formato",
        "cantidad",
        "rango",
        "caras",
        "terminacion",
        "adicional",
        "modo_precio",
        "precio_final",
        "precio_unitario",
        "componente",
        "tipo_componente",
        "variable_key",
        "variable_label",
        "valor_base",
        "unidad",
        "operacion",
        "factor_multiplicador",
        "subtotal_componente",
        "fuente_componente",
        "fuente_precio_final",
        "editable_en_sistema",
        "editable_en_excel_maestro",
        "impacta_hoy",
        "estado_operativo",
        "observacion",
    ]

    def __init__(
        self,
        project_root: Path,
        validation_excel_status: str = "ok",
        validation_report_path: str | None = None,
    ):
        self.project_root = project_root
        self.variables_service = PrincipalVariablesService(project_root)
        self.generated_at = datetime.now()
        self.validation_excel_status = validation_excel_status
        self.validation_report_path = validation_report_path

    def render(self, export: dict[str, Any]) -> tuple[str, bytes]:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for sheet_name in self.REQUIRED_SHEETS:
            workbook.create_sheet(sheet_name)

        self._build_summary(workbook["00_RESUMEN"], export)
        self._build_variables(workbook["01_VARIABLES_MADRE"])
        self._build_ranges(workbook["02_RANGOS"])
        self._build_product_sheets(workbook, export)
        self._build_blocked(workbook["18_BLOQUEADOS"], export)
        self._build_trace(workbook["19_TRAZABILIDAD"], export)
        self._build_price_trace(workbook["21_TRAZABILIDAD_PRECIOS"], export)

        self._sanitize_workbook_text(workbook)
        for worksheet in workbook.worksheets:
            self._format_sheet(worksheet)

        buffer = BytesIO()
        workbook.save(buffer)
        payload = buffer.getvalue()
        filename = f"cotizador_digital_maestro_{self.generated_at:%Y%m%d_%H%M%S}.xlsx"
        self._save_report_copy(filename, payload)
        return filename, payload

    def _build_summary(self, ws, export: dict[str, Any]) -> None:
        tables = export.get("tablas", [])
        blocked = self._blocked_rows(export)
        variables = self._mother_variables()
        version = self._git_version_info()
        operative = [item for item in variables if self._is_operative_variable(item)]
        prepared = [item for item in variables if not self._is_operative_variable(item)]
        ws.append(["Campo", "Valor"])
        rows = [
            ("fecha_generacion", self.generated_at.isoformat(timespec="seconds")),
            ("version_base_git", version["version_base_git"]),
            ("commit_actual", version["commit_actual"]),
            ("repo_dirty", version["repo_dirty"]),
            ("nota_version", version["nota_version"]),
            ("fuente_precios_finales", "lista-low.pdf"),
            ("fuente_logica_historica", "Copia de DIGITAL sistema 2025.xlsx"),
            ("precios_sin_iva", True),
            ("tablas_pdf_fijas_no_editables", True),
            ("variables_madre_operativas", len(operative)),
            ("variables_madre_preparadas", len(prepared)),
            ("productos_activos", len([table for table in tables if table.get("estado") == "activo"])),
            ("productos_bloqueados", len(blocked)),
            ("hojas_generadas", len(self.REQUIRED_SHEETS)),
            ("validacion_excel", self.validation_excel_status),
            ("reporte_validacion", self.validation_report_path or "ver tests/api/test_variables_principales_api.py y scripts/qa/validar_excel_maestro.py"),
            ("trazabilidad_precios", True),
            ("hoja_trazabilidad_precios", "21_TRAZABILIDAD_PRECIOS"),
            ("nota", "Productos PDF fijo no se editan desde este Excel; las variables madre se auditan por separado."),
        ]
        for row in rows:
            ws.append(row)

    def _build_variables(self, ws) -> None:
        headers = [
            "key",
            "label",
            "value",
            "unit",
            "editable",
            "editable_en_sistema",
            "editable_en_excel_maestro",
            "tipo",
            "impacta_hoy",
            "estado_operativo",
            "productos_afectados",
            "source_file",
            "source_path",
            "source_sheet",
            "source_cell",
            "description",
        ]
        ws.append(headers)
        forbidden = ("precio_final", "precio_pdf", "precio_sin_iva", "precio_objetivo_pdf", "factor_ajuste_pdf", "total_final", "matriz_pdf")
        for item in self._mother_variables():
            key = str(item.get("key", ""))
            if key == "rango" or any(token in key for token in forbidden):
                continue
            impacts_today = self._is_operative_variable(item)
            editable_in_system = impacts_today
            editable_in_workbook = True
            estado_operativo = "operativa" if impacts_today else "preparada_no_conectada"
            source_sheet, source_cell = self._split_source_cell(item.get("source_cell"))
            source_path = item.get("source_path")
            if not source_path and isinstance(item.get("path"), str):
                source_path = item.get("path")
            description = item.get("description")
            if not impacts_today:
                description = (
                    "Costo base detectado en Excel histórico. Preparado para fórmulas variables futuras. "
                    "Cambiar este valor en el Excel maestro no modifica hoy los precios finales del cotizador."
                )
            ws.append([
                key,
                item.get("label"),
                item.get("value"),
                item.get("unit"),
                editable_in_system,
                editable_in_system,
                editable_in_workbook,
                item.get("tipo"),
                impacts_today,
                estado_operativo,
                ", ".join(item.get("productos_afectados") or []),
                item.get("source_file"),
                source_path,
                source_sheet,
                source_cell,
                description,
            ])

    def _build_ranges(self, ws) -> None:
        ws.append(["grupo", "rango", "orden", "editable", "tipo", "fuente", "motivo"])
        ranges = self.variables_service.ranges().get("rangos", [])
        for entry in ranges:
            for index, rango in enumerate(entry.get("rangos", []), start=1):
                ws.append([
                    entry.get("grupo"),
                    rango,
                    index,
                    False,
                    entry.get("tipo", "rango_fijo"),
                    entry.get("fuente"),
                    entry.get("motivo"),
                ])

    def _build_product_sheets(self, workbook: Workbook, export: dict[str, Any]) -> None:
        by_title = {table.get("titulo"): table for table in export.get("tablas", [])}
        bajadas = by_title.get("Bajadas Fullcolor / ByN / Kraft", {})
        self._write_product_table(
            workbook["03_BAJADAS_FULLCOLOR"],
            "Bajadas Fullcolor / Kraft",
            self._filter_table_rows(bajadas, lambda row: row.get("categoria") != "Bajadas Blanco y Negro"),
            bajadas.get("fuente"),
            "tabla_fija_pdf + adicionales separados",
        )
        self._write_product_table(
            workbook["04_BAJADAS_BYN"],
            "Bajadas Blanco y Negro",
            self._filter_table_rows(bajadas, lambda row: row.get("categoria") == "Bajadas Blanco y Negro"),
            bajadas.get("fuente"),
            "tabla_fija_pdf",
        )

        for title, sheet_name in self.SHEET_BY_TITLE.items():
            if sheet_name in {"03_BAJADAS_FULLCOLOR", "04_BAJADAS_BYN"}:
                continue
            table = by_title.get(title)
            if not table:
                continue
            self._write_product_table(
                workbook[sheet_name],
                title,
                self._rows_as_dicts(table),
                table.get("fuente"),
                self._product_mode(title),
            )

        self._append_autoadhesivas_additionals(workbook["05_AUTOADHESIVAS"])

    def _write_product_table(self, ws, title: str, rows: list[dict[str, Any]], source: str | None, mode: str) -> None:
        ws.append(["Producto", title])
        ws.append(["Estado", "activo"])
        ws.append(["Fuente", source or "matriz interna"])
        ws.append(["Tipo", mode])
        ws.append(["Precios sin IVA", "si"])
        ws.append(["Observaciones", "Tabla final vigente; no editar precios finales manualmente."])
        ws.append([])
        headers = self._columns_for_rows(rows)
        ws.append(headers or ["sin_datos"])
        for row in rows:
            ws.append([row.get(header) for header in headers])

    def _append_autoadhesivas_additionals(self, ws) -> None:
        config = self._read_json("data/bajadas_autoadhesivas/autoadhesivas_v1_config.json")
        tinta_base = config.get("adicional_tinta_blanca_base_1_copia")
        start = ws.max_row + 3
        ws.cell(start, 1, "Adicionales Autoadhesivas")
        ws.cell(start, 1).font = Font(bold=True)
        ws.append(["adicional", "rango", "precio_unitario_sin_iva", "fuente", "regla"])
        for rango, value in self.LACA_UV_BAJADAS.items():
            ws.append(["laca_uv", rango, value, "matriz_laca_uv_bajadas", "unitario_por_rango_x_cantidad"])
        ws.append([
            "tinta_blanca",
            "1 copia proporcional",
            tinta_base,
            "data/bajadas_autoadhesivas/autoadhesivas_v1_config.json#adicional_tinta_blanca_base_1_copia",
            "valor_base_1_copia_x_cantidad",
        ])

    def _build_blocked(self, ws, export: dict[str, Any]) -> None:
        ws.append(["producto", "estado", "motivo", "fuente", "accion_necesaria", "precio_inventado"])
        for row in self._blocked_rows(export):
            ws.append([
                row.get("producto"),
                row.get("estado"),
                row.get("motivo"),
                row.get("fuente"),
                row.get("accion_necesaria"),
                False,
            ])

    def _build_trace(self, ws, export: dict[str, Any]) -> None:
        ws.append([
            "producto",
            "endpoint",
            "modo_precio",
            "fuente_pdf",
            "fuente_excel",
            "archivo_data",
            "motor_backend",
            "estado",
            "notas",
        ])
        endpoints = {
            "Bajadas Fullcolor / ByN / Kraft": "/bajadas-v2/cotizar",
            "Bajadas Autoadhesivas": "/bajadas-v2/cotizar",
            "Stickers Circulares": "/stickers-circulares/cotizar",
            "Troquelado Digital": "adicional en /bajadas-v2/cotizar",
            "Tarjetas Troqueladas Circulares": "/tarjetas-troqueladas-circulares/cotizar",
            "Tarjetas Personales": "/tarjetas-9x5/cotizar",
            "Tarjetas Postales": "/tarjetas-postales/cotizar",
            "Folletos": "/folletos/cotizar",
            "Carpetas": "/carpetas/cotizar",
            "Sobres": "/sobres/cotizar",
            "Stickers Corte Recto": "/stickers-corte-recto/cotizar",
            "Imanes Corte Recto": "/imanes-corte-recto/cotizar",
            "Plancha de Imán Impreso": "/plancha-iman-impreso/cotizar",
            "Plancha de Im?n Impreso": "/plancha-iman-impreso/cotizar",
            "Agendas / Cuadernos": "/agendas-cuadernos/cotizar",
        }
        motors = {
            "Bajadas Fullcolor / ByN / Kraft": "src/bajadas_v2/pricing_engine.py",
            "Bajadas Autoadhesivas": "src/bajadas_autoadhesivas/pricing_engine.py",
            "Stickers Circulares": "src/stickers_circulares/pricing_engine.py",
        }
        for table in export.get("tablas", []):
            title = table.get("titulo")
            if table.get("estado") != "activo":
                continue
            ws.append([
                title,
                endpoints.get(title, "-"),
                self._product_mode(title),
                self._pdf_source_for(title),
                self._excel_source_for(title),
                table.get("fuente"),
                motors.get(title, f"src/{self._slug(title)}/pricing_engine.py"),
                "activo",
                "PDF es fuente final; Excel aporta lógica histórica/variables cuando existe.",
            ])
        for row in self._blocked_rows(export):
            ws.append([
                row.get("producto"),
                "-",
                "bloqueado",
                row.get("fuente"),
                "Copia de DIGITAL sistema 2025.xlsx",
                row.get("fuente"),
                "-",
                row.get("estado"),
                row.get("motivo"),
            ])

    def _build_price_trace(self, ws, export: dict[str, Any]) -> None:
        ws.append(self.PRICE_TRACE_HEADERS)
        variable_by_key = {str(item.get("key")): item for item in self._mother_variables()}
        auto_config = self._read_json("data/bajadas_autoadhesivas/autoadhesivas_v1_config.json")
        circular_config = merge_global_base_costs(
            self._read_json("data/stickers_circulares/formula_editable_config.json"),
            self.project_root,
        )
        circular_factor = self._find_circular_factor("obra_ilustracion_90g", "10cm", "con_laca_uv", 1000)
        tinta_base = auto_config.get("adicional_tinta_blanca_base_1_copia")

        rows = [
            self._price_trace_row(
                producto="Bajadas Kraft",
                familia="Bajadas",
                material="Kraft",
                gramaje="80g",
                formato="A3",
                cantidad=1,
                rango="1",
                caras="4/0",
                modo_precio="tabla_fija_pdf",
                precio_final=782,
                precio_unitario=782,
                componente="precio_base_pdf",
                tipo_componente="tabla_pdf",
                operacion="precio tomado directo de matriz PDF",
                subtotal_componente=782,
                fuente_componente="data/bajadas_v2/precios_pdf_objetivo_limpio.json",
                fuente_precio_final="lista-low.pdf",
                estado_operativo="tabla_pdf_fija",
                observacion="Kraft A3 usa matriz PDF específica; no se recalcula desde variables madre.",
            ),
            self._price_trace_row(
                producto="Bajadas Kraft",
                familia="Bajadas",
                material="Kraft",
                gramaje="80g",
                formato="A3",
                cantidad=100,
                rango="51 a 100",
                caras="4/0",
                modo_precio="tabla_fija_pdf",
                precio_final=56900,
                precio_unitario=569,
                componente="precio_base_pdf",
                tipo_componente="tabla_pdf",
                operacion="precio unitario PDF x cantidad",
                subtotal_componente=56900,
                fuente_componente="data/bajadas_v2/precios_pdf_objetivo_limpio.json",
                fuente_precio_final="lista-low.pdf",
                estado_operativo="tabla_pdf_fija",
                observacion="Precio final formado por unitario 569 por 100 unidades.",
            ),
            self._price_trace_row(
                producto="Bajadas Fullcolor",
                familia="Bajadas",
                material="Ilustracion",
                gramaje="150g",
                formato="A3+",
                cantidad=53,
                rango="51 a 100",
                caras="4/0",
                modo_precio="tabla_pdf_fija_con_adicional_variable",
                precio_final=32966,
                precio_unitario=622,
                componente="precio_base_pdf",
                tipo_componente="tabla_pdf",
                variable_key="ilustracion_150g_65x95_usd",
                variable_label="Ilustración 150g 65x95",
                operacion="precio tomado directo de matriz PDF",
                subtotal_componente=32966,
                fuente_componente="data/bajadas_v2/precios_pdf_objetivo_limpio.json",
                fuente_precio_final="lista-low.pdf",
                editable_en_excel_maestro=True,
                estado_operativo="tabla_pdf_fija",
                observacion="El papel está relacionado conceptualmente, pero cambiar la variable preparada no modifica hoy Bajadas Fullcolor.",
            ),
            self._price_trace_row(
                producto="Bajadas Fullcolor",
                familia="Bajadas",
                material="Ilustracion",
                gramaje="150g",
                formato="A3+",
                cantidad=53,
                rango="51 a 100",
                caras="4/0",
                adicional="laca_uv",
                modo_precio="tabla_pdf_fija_con_adicional_variable",
                precio_final=38584,
                precio_unitario=728,
                componente="laca_uv",
                tipo_componente="adicional_variable",
                valor_base=106,
                unidad="ARS/unidad",
                operacion="valor x cantidad",
                subtotal_componente=5618,
                fuente_componente="matriz_laca_uv_bajadas",
                fuente_precio_final="lista-low.pdf + matriz_laca_uv_bajadas",
                impacta_hoy=True,
                estado_operativo="adicional_operativo",
                observacion="Laca UV usa escala fija por rango validada; en rango 51 a 100 vale 106 por unidad.",
            ),
            self._price_trace_row(
                producto="Bajadas Autoadhesivas",
                familia="Bajadas",
                variante="especial",
                material="OPP blanco",
                formato="A3+",
                cantidad=30,
                rango="26 a 50",
                caras="4/0",
                adicional="laca_uv",
                modo_precio="tabla_pdf_fija_con_adicional_variable",
                componente="laca_uv",
                tipo_componente="adicional_variable",
                valor_base=116,
                unidad="ARS/unidad",
                operacion="valor x cantidad",
                subtotal_componente=3480,
                fuente_componente="matriz_laca_uv_bajadas",
                fuente_precio_final="lista-low.pdf + matriz_laca_uv_bajadas",
                impacta_hoy=True,
                estado_operativo="adicional_operativo",
                observacion="Laca UV en autoadhesivas se suma como adicional acumulable; no duplica con payload legacy.",
            ),
            self._price_trace_row(
                producto="Bajadas Autoadhesivas",
                familia="Bajadas",
                variante="especial",
                material="OPP blanco",
                formato="A3+",
                cantidad=30,
                rango="26 a 50",
                caras="4/0",
                adicional="tinta_blanca",
                modo_precio="tabla_pdf_fija_con_adicional_variable",
                componente="tinta_blanca",
                tipo_componente="adicional_variable",
                variable_key="adicional_tinta_blanca_base_1_copia",
                variable_label=self._variable_label(variable_by_key, "adicional_tinta_blanca_base_1_copia", "Tinta blanca Autoadhesivas (1 copia)"),
                valor_base=tinta_base,
                unidad="ARS/unidad",
                operacion="valor_base x cantidad",
                subtotal_componente=tinta_base * 30 if isinstance(tinta_base, (int, float)) else None,
                fuente_componente="data/bajadas_autoadhesivas/autoadhesivas_v1_config.json",
                fuente_precio_final="lista-low.pdf + adicional operativo",
                editable_en_sistema=True,
                editable_en_excel_maestro=True,
                impacta_hoy=True,
                estado_operativo="adicional_operativo",
                observacion="La base de tinta blanca impacta hoy como adicional proporcional por cantidad.",
            ),
            self._price_trace_row(
                producto="Tarjetas 9x5",
                familia="Tarjetas",
                material="Ilustracion",
                gramaje="300g",
                formato="9x5",
                cantidad=100,
                caras="4/0",
                terminacion="sin_laminar",
                modo_precio="tabla_fija_pdf",
                precio_final=5139,
                precio_unitario=51.39,
                componente="precio_base_pdf",
                tipo_componente="tabla_pdf",
                operacion="precio tomado directo de matriz PDF",
                subtotal_componente=5139,
                fuente_componente="data/tarjetas_9x5/precios_pdf_objetivo_tarjetas_9x5.json",
                fuente_precio_final="lista-low.pdf",
                estado_operativo="tabla_pdf_fija",
                observacion="Precio final cerrado por matriz PDF. No se edita desde variables madre.",
            ),
            self._price_trace_row(
                producto="Tarjetas 9x5",
                familia="Tarjetas",
                material="Ilustracion",
                gramaje="300g",
                formato="9x5",
                cantidad=1000,
                caras="4/4",
                terminacion="laminado_mate",
                modo_precio="tabla_fija_pdf",
                precio_final=48401,
                precio_unitario=48.401,
                componente="precio_base_pdf",
                tipo_componente="tabla_pdf",
                operacion="precio tomado directo de matriz PDF",
                subtotal_componente=48401,
                fuente_componente="data/tarjetas_9x5/precios_pdf_objetivo_tarjetas_9x5.json",
                fuente_precio_final="lista-low.pdf",
                estado_operativo="tabla_pdf_fija",
                observacion="Terminación incluida en matriz final PDF.",
            ),
            self._price_trace_row(
                producto="Tarjetas Postales",
                familia="Tarjetas",
                material="Ilustracion",
                gramaje="300g",
                formato="postal",
                cantidad=100,
                caras="4/0",
                terminacion="sin_laminar",
                modo_precio="tabla_fija_pdf",
                precio_final=10932,
                precio_unitario=109.32,
                componente="precio_base_pdf",
                tipo_componente="tabla_pdf",
                operacion="precio tomado directo de matriz PDF",
                subtotal_componente=10932,
                fuente_componente="data/tarjetas_postales/precios_pdf_objetivo_tarjetas_postales.json",
                fuente_precio_final="lista-low.pdf",
                estado_operativo="tabla_pdf_fija",
                observacion="Precio final cerrado por matriz PDF.",
            ),
            self._price_trace_row(
                producto="Tarjetas Postales",
                familia="Tarjetas",
                material="Ilustracion",
                gramaje="300g",
                formato="postal",
                cantidad=1000,
                caras="4/4",
                terminacion="laminado_mate",
                modo_precio="tabla_fija_pdf",
                precio_final=136795,
                precio_unitario=136.795,
                componente="precio_base_pdf",
                tipo_componente="tabla_pdf",
                operacion="precio tomado directo de matriz PDF",
                subtotal_componente=136795,
                fuente_componente="data/tarjetas_postales/precios_pdf_objetivo_tarjetas_postales.json",
                fuente_precio_final="lista-low.pdf",
                estado_operativo="tabla_pdf_fija",
                observacion="Terminación incluida en matriz final PDF.",
            ),
        ]

        rows.extend(self._fixed_price_trace_rows())
        rows.extend(self._stickers_circulares_trace_rows(variable_by_key, circular_config, circular_factor))
        rows.extend(self._blocked_price_trace_rows(export))
        for row in rows:
            ws.append([row.get(header) for header in self.PRICE_TRACE_HEADERS])

    def _price_trace_row(self, **kwargs: Any) -> dict[str, Any]:
        row = {header: None for header in self.PRICE_TRACE_HEADERS}
        row.update({
            "editable_en_sistema": False,
            "editable_en_excel_maestro": False,
            "impacta_hoy": False,
            "estado_operativo": "tabla_pdf_fija",
        })
        row.update(kwargs)
        return row

    def _fixed_price_trace_rows(self) -> list[dict[str, Any]]:
        specs = [
            ("Folletos", "Folletos", "Ilustracion", "80g", "A4", 1000, None, "1/1", None, None, 119247, 119.247, "data/folletos/precios_pdf_objetivo_folletos.json", "A4 80g escala grises 1/1."),
            ("Folletos", "Folletos", "Ilustracion", "150g", "10x15", 100, None, "4/0", None, None, 8445, 84.45, "data/folletos/precios_pdf_objetivo_folletos.json", "10x15 150g fullcolor 4/0."),
            ("Folletos", "Folletos", "Ilustracion", "150g", "10x10", 100, None, "4/0", None, None, 6496, 64.96, "data/folletos/precios_pdf_objetivo_folletos.json", "10x10 150g fullcolor 4/0."),
            ("Carpetas", "Papelería", "Ilustracion", "300g", "A4", 1, "1", "4/0", "sin_laminar", None, 1762, 1762, "data/carpetas/precios_pdf_objetivo_carpetas.json", "Sin solapa impresa."),
            ("Carpetas", "Papelería", "Ilustracion", "300g", "A4", 1, "1", "4/0", "sin_laminar", "solapa_impresa", 2017, 2017, "data/carpetas/precios_pdf_objetivo_carpetas.json", "Base 1762 + solapa 255."),
            ("Carpetas", "Papelería", "Ilustracion", "300g", "A4", 100, "51 a 100", "4/4", "laca_uv", "solapa_impresa", 187100, 1871, "data/carpetas/precios_pdf_objetivo_carpetas.json", "Unitario 1683 + solapa 188, multiplicado por 100."),
            ("Sobres", "Papelería", "Papel blanco", "63g", "sobre_bolsa_a4_22_9x32_4", 100, None, "4/0", None, None, 63300, 633, "data/sobres/precios_pdf_objetivo_sobres.json", "Precio unitario PDF por cantidad."),
            ("Sobres", "Papelería", "Papel blanco", "63g", "sobre_bolsa_27x37", 1000, None, "4/0", None, None, 536000, 536, "data/sobres/precios_pdf_objetivo_sobres.json", "Precio unitario PDF por cantidad."),
            ("Stickers Corte Recto", "Stickers", "Sticker", None, "6x4", 100, None, "4/0", "sin_laca_uv", None, 2765, 27.65, "data/stickers_corte_recto/precios_pdf_objetivo_stickers_corte_recto.json", "Precio total por paquete PDF."),
            ("Stickers Corte Recto", "Stickers", "Sticker", None, "10x7", 1000, None, "4/0", "con_laca_uv", None, 61703, 61.703, "data/stickers_corte_recto/precios_pdf_objetivo_stickers_corte_recto.json", "Precio total por paquete PDF."),
            ("Imanes Corte Recto", "Imanes", "Imán + Ilustracion", "300g", "6x4", 100, None, "4/0", "sin_laca_uv", None, 7526, 75.26, "data/imanes_corte_recto/precios_pdf_objetivo_imanes_corte_recto.json", "Precio total por paquete PDF."),
            ("Imanes Corte Recto", "Imanes", "Imán + Ilustracion", "300g", "10x7", 1000, None, "4/0", "con_laca_uv", None, 153680, 153.68, "data/imanes_corte_recto/precios_pdf_objetivo_imanes_corte_recto.json", "Precio total por paquete PDF."),
            ("Plancha de Imán Impreso", "Imanes", "Papel 300g Ilustracion", "300g", "30x46", 500, "301 a 500", "4/0", None, None, 965500, 1931, "data/plancha_iman_impreso/precios_pdf_objetivo_plancha_iman_impreso.json", "Unitario 1931 por 500."),
            ("Agendas / Cuadernos", "Agendas", "Agenda 2026", None, "A5", 2, None, None, None, None, 6000, 3000, "data/agendas_cuadernos/precios_pdf_objetivo_agendas_cuadernos.json", "Agenda 2026 A5 72 páginas, precio unitario test actual."),
        ]
        return [
            self._price_trace_row(
                producto=producto,
                familia=familia,
                material=material,
                gramaje=gramaje,
                formato=formato,
                cantidad=cantidad,
                rango=rango,
                caras=caras,
                terminacion=terminacion,
                adicional=adicional,
                modo_precio="tabla_fija_pdf",
                precio_final=precio_final,
                precio_unitario=precio_unitario,
                componente="precio_base_pdf",
                tipo_componente="tabla_pdf",
                operacion="precio tomado directo de matriz PDF",
                subtotal_componente=precio_final,
                fuente_componente=fuente,
                fuente_precio_final="lista-low.pdf",
                estado_operativo="tabla_pdf_fija",
                observacion=observacion,
            )
            for producto, familia, material, gramaje, formato, cantidad, rango, caras, terminacion, adicional, precio_final, precio_unitario, fuente, observacion in specs
        ]

    def _stickers_circulares_trace_rows(
        self,
        variable_by_key: dict[str, dict[str, Any]],
        config: dict[str, Any],
        factor_row: dict[str, Any] | None,
    ) -> list[dict[str, Any]]:
        variables = config.get("variables", {})
        coef_tamano = (variables.get("coeficiente_tamano") or {}).get("10cm")
        coef_cantidad = (variables.get("coeficiente_cantidad") or {}).get("1000")
        material_base = (variables.get("material_base") or {}).get("obra_ilustracion_90g")
        click_color = variables.get("click_color_base")
        laca_factor = variables.get("laca_uv_factor")
        corte_factor = variables.get("corte_circular_factor")
        multiplicador = variables.get("multiplicador_comercial")
        precio_base = factor_row.get("precio_base_excel") if factor_row else None
        factor_ajuste = factor_row.get("factor_ajuste_pdf") if factor_row else None
        precio_pdf = factor_row.get("precio_pdf_objetivo") if factor_row else 85980
        base_common = {
            "producto": "Stickers Circulares",
            "familia": "Stickers",
            "variante": "circular",
            "material": "obra_ilustracion_90g",
            "gramaje": "90g",
            "formato": "10cm",
            "cantidad": 1000,
            "caras": "4/0",
            "terminacion": "con_laca_uv",
            "modo_precio": "formula_editable_calibrada",
            "precio_final": precio_pdf,
            "precio_unitario": round(precio_pdf / 1000, 4) if isinstance(precio_pdf, (int, float)) else None,
            "fuente_precio_final": "fórmula editable calibrada contra lista-low.pdf",
        }
        return [
            self._price_trace_row(
                **base_common,
                componente="papel",
                tipo_componente="variable_madre",
                variable_key="obra_90g",
                variable_label=self._variable_label(variable_by_key, "obra_90g", "Papel obra/ilustración 90g"),
                valor_base=material_base,
                unidad="USD",
                operacion="participa en subtotal_formula_excel",
                fuente_componente="data/variables_globales/costos_base.json",
                editable_en_sistema=True,
                editable_en_excel_maestro=True,
                impacta_hoy=True,
                estado_operativo="operativa",
                observacion="El costo de papel impacta hoy en la fórmula editable calibrada de Stickers Circulares.",
            ),
            self._price_trace_row(
                **base_common,
                componente="click_color",
                tipo_componente="variable_madre",
                variable_key="click_color",
                variable_label=self._variable_label(variable_by_key, "click_color", "Click color"),
                valor_base=click_color,
                unidad="ARS",
                operacion="participa en subtotal_formula_excel",
                fuente_componente="data/stickers_circulares/formula_editable_config.json",
                editable_en_sistema=True,
                editable_en_excel_maestro=True,
                impacta_hoy=True,
                estado_operativo="operativa",
                observacion="Click color operativo usado por fórmula calibrada.",
            ),
            self._price_trace_row(
                **base_common,
                componente="laca_uv",
                tipo_componente="factor",
                valor_base=laca_factor,
                unidad="factor",
                operacion="subtotal x factor",
                factor_multiplicador=laca_factor,
                fuente_componente="data/stickers_circulares/formula_editable_config.json",
                editable_en_excel_maestro=True,
                impacta_hoy=True,
                estado_operativo="derivada",
                observacion="Factor de laca UV de la fórmula editable calibrada.",
            ),
            self._price_trace_row(
                **base_common,
                componente="corte_circular",
                tipo_componente="factor",
                valor_base=corte_factor,
                unidad="factor",
                operacion="subtotal x factor",
                factor_multiplicador=corte_factor,
                fuente_componente="data/stickers_circulares/formula_editable_config.json",
                editable_en_excel_maestro=True,
                impacta_hoy=True,
                estado_operativo="derivada",
                observacion="Factor de corte circular detectado en configuración editable.",
            ),
            self._price_trace_row(
                **base_common,
                componente="coeficiente_tamano",
                tipo_componente="factor",
                valor_base=coef_tamano,
                unidad="factor",
                operacion="valor x coeficiente",
                factor_multiplicador=coef_tamano,
                fuente_componente="data/stickers_circulares/formula_editable_config.json",
                editable_en_excel_maestro=True,
                impacta_hoy=True,
                estado_operativo="derivada",
                observacion="Coeficiente de tamaño 10cm.",
            ),
            self._price_trace_row(
                **base_common,
                componente="coeficiente_cantidad",
                tipo_componente="factor",
                valor_base=coef_cantidad,
                unidad="factor",
                operacion="valor x coeficiente",
                factor_multiplicador=coef_cantidad,
                fuente_componente="data/stickers_circulares/formula_editable_config.json",
                editable_en_excel_maestro=True,
                impacta_hoy=True,
                estado_operativo="derivada",
                observacion="Coeficiente de cantidad 1000.",
            ),
            self._price_trace_row(
                **base_common,
                componente="multiplicador_general",
                tipo_componente="multiplicador",
                variable_key="multiplicador_general",
                variable_label=self._variable_label(variable_by_key, "multiplicador_general", "Multiplicador comercial general"),
                valor_base=multiplicador,
                unidad="factor",
                operacion="subtotal x factor",
                factor_multiplicador=multiplicador,
                fuente_componente="data/stickers_circulares/formula_editable_config.json",
                editable_en_sistema=True,
                editable_en_excel_maestro=True,
                impacta_hoy=True,
                estado_operativo="operativa",
                observacion="Multiplicador operativo de la fórmula editable calibrada.",
            ),
            self._price_trace_row(
                **base_common,
                componente="subtotal_formula",
                tipo_componente="derivado",
                valor_base=precio_base,
                unidad="ARS",
                operacion="subtotal reconstruido desde fórmula Excel",
                subtotal_componente=precio_base,
                fuente_componente="data/stickers_circulares/factores_ajuste_pdf.json",
                estado_operativo="derivada",
                observacion="Subtotal base antes de calibración PDF.",
            ),
            self._price_trace_row(
                **base_common,
                componente="factor_ajuste_pdf",
                tipo_componente="factor",
                valor_base=factor_ajuste,
                unidad="factor",
                operacion="subtotal x factor",
                factor_multiplicador=factor_ajuste,
                subtotal_componente=precio_pdf,
                fuente_componente="data/stickers_circulares/factores_ajuste_pdf.json",
                estado_operativo="derivada",
                observacion="Factor de ajuste calibra la fórmula para coincidir con PDF; no es variable madre editable.",
            ),
            self._price_trace_row(
                **base_common,
                componente="total_final",
                tipo_componente="derivado",
                valor_base=precio_pdf,
                unidad="ARS",
                operacion="subtotal_formula x factor_ajuste_pdf",
                subtotal_componente=precio_pdf,
                fuente_componente="data/stickers_circulares/precios_pdf_objetivo_stickers_circulares.json",
                estado_operativo="derivada",
                observacion="Total final coincide con PDF vigente.",
            ),
        ]

    def _blocked_price_trace_rows(self, export: dict[str, Any]) -> list[dict[str, Any]]:
        rows = []
        for blocked in self._blocked_rows(export):
            rows.append(self._price_trace_row(
                producto=blocked.get("producto"),
                familia="Bloqueados",
                modo_precio="bloqueado",
                componente="bloqueado",
                tipo_componente="bloqueado",
                operacion="bloqueado por falta de datos",
                fuente_componente=blocked.get("fuente"),
                fuente_precio_final="bloqueado",
                estado_operativo="bloqueada",
                observacion=blocked.get("motivo"),
            ))
        return rows

    def _variable_label(self, variable_by_key: dict[str, dict[str, Any]], key: str, fallback: str) -> str:
        return str(variable_by_key.get(key, {}).get("label") or fallback)

    def _find_circular_factor(self, material: str, formato: str, terminacion: str, cantidad: int) -> dict[str, Any] | None:
        document = self._read_json("data/stickers_circulares/factores_ajuste_pdf.json")
        rows = document.get("rows", document) if isinstance(document, dict) else document
        if not isinstance(rows, list):
            return None
        for row in rows:
            if not isinstance(row, dict):
                continue
            if (
                row.get("material") == material
                and row.get("formato") == formato
                and row.get("terminacion") == terminacion
                and row.get("cantidad_unidades") == cantidad
            ):
                return row
        return None

    def _format_sheet(self, ws) -> None:
        thin = Side(style="thin", color="D9E2EC")
        border = Border(bottom=thin)
        header_fill = PatternFill("solid", fgColor="D8EAFE")
        editable_fill = PatternFill("solid", fgColor="D9F99D")
        fixed_fill = PatternFill("solid", fgColor="E5E7EB")
        derived_fill = PatternFill("solid", fgColor="DBEAFE")
        blocked_fill = PatternFill("solid", fgColor="FED7AA")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row == 1 or self._looks_like_header(cell):
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.border = border
                if isinstance(cell.value, (int, float)) and "precio" in str(ws.cell(cell.row, 1).value).lower():
                    cell.number_format = '#,##0.00'
                if isinstance(cell.value, str):
                    lowered = cell.value.lower()
                    if lowered in {"true", "variable_madre", "operativa"}:
                        cell.fill = editable_fill
                    elif "pdf fijo" in lowered or "tabla fija" in lowered or "tabla_fija_pdf" in lowered:
                        cell.fill = fixed_fill
                    elif "derivado" in lowered or "formula" in lowered or "fórmula" in lowered or "preparada_no_conectada" in lowered:
                        cell.fill = derived_fill
                    elif "bloqueado" in lowered or "excluido" in lowered:
                        cell.fill = blocked_fill
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        for column in ws.columns:
            max_len = 0
            letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)

    def _mother_variables(self) -> list[dict[str, Any]]:
        grouped = self.variables_service.get_grouped()
        result = []
        for group in PrincipalVariablesService.GROUP_ORDER:
            result.extend(grouped.get(group, []))
        return result

    def _is_operative_variable(self, item: dict[str, Any]) -> bool:
        return bool(item.get("impacta_hoy")) or item.get("key") == "tipo_cambio_usd"

    def _blocked_rows(self, export: dict[str, Any]) -> list[dict[str, Any]]:
        rows = []
        for table in export.get("tablas", []):
            if table.get("estado") == "bloqueado":
                rows.append({
                    "producto": table.get("titulo"),
                    "estado": table.get("estado"),
                    "motivo": table.get("motivo"),
                    "accion_necesaria": "Aportar datos finales confiables y convención de precio.",
                    "fuente": table.get("fuente"),
                })
        rows.extend(self.BLOCKED_EXTRA)
        return rows

    def _rows_as_dicts(self, table: dict[str, Any]) -> list[dict[str, Any]]:
        columns = table.get("columnas") or []
        return [dict(zip(columns, row)) for row in table.get("filas", [])]

    def _filter_table_rows(self, table: dict[str, Any], predicate) -> list[dict[str, Any]]:
        return [row for row in self._rows_as_dicts(table) if predicate(row)]

    def _columns_for_rows(self, rows: list[dict[str, Any]]) -> list[str]:
        columns: list[str] = []
        for row in rows:
            for key in row:
                if key not in columns:
                    columns.append(key)
        return columns

    def _product_mode(self, title: str | None) -> str:
        if title in {"Stickers Circulares", "Stickers Corte Recto", "Imanes Corte Recto"}:
            return "formula_editable_calibrada"
        if title in {
            "Bajadas Fullcolor / ByN / Kraft",
            "Bajadas Autoadhesivas",
            "Tarjetas Personales",
            "Tarjetas Postales",
            "Folletos",
            "Carpetas",
            "Sobres",
            "Plancha de Im?n Impreso",
            "Plancha de Im?n Impreso",
            "Agendas / Cuadernos",
        }:
            return "matriz_pdf_con_variables_contextuales"
        if title == "Troquelado Digital":
            return "variable"
        return "tabla_fija_pdf"

    def _pdf_source_for(self, title: str | None) -> str:
        sources = {
            "Bajadas Fullcolor / ByN / Kraft": "PDF páginas 4 a 6",
            "Bajadas Autoadhesivas": "PDF página 7",
            "Stickers Circulares": "PDF hoja 8",
            "Troquelado Digital": "PDF página 11",
            "Tarjetas Troqueladas Circulares": "PDF página 11",
            "Tarjetas Personales": "PDF página 12",
            "Tarjetas Postales": "PDF página 12",
            "Folletos": "PDF página 13",
            "Carpetas": "PDF página 14",
            "Sobres": "PDF página 14",
            "Stickers Corte Recto": "PDF página 15",
            "Imanes Corte Recto": "PDF página 15",
            "Plancha de Imán Impreso": "PDF página 16",
            "Plancha de Im?n Impreso": "PDF página 16",
            "Agendas / Cuadernos": "PDF página 17",
        }
        return sources.get(title, "lista-low.pdf")

    def _excel_source_for(self, title: str | None) -> str:
        if title == "Bajadas Fullcolor / ByN / Kraft":
            return "Excel histórico + data/bajadas_v2/formula_editable_config.json"
        if title == "Bajadas Autoadhesivas":
            return "Excel histórico + data/bajadas_v2/formula_editable_config.json + data/bajadas_autoadhesivas/autoadhesivas_v1_config.json"
        if title == "Stickers Circulares":
            return 'Excel hoja "circulares" + data/stickers_circulares/formula_editable_config.json'
        if title == "Stickers Corte Recto":
            return "Excel histórico + data/stickers_corte_recto/formula_editable_config.json"
        if title == "Imanes Corte Recto":
            return "Excel histórico + data/imanes_corte_recto/formula_editable_config.json"
        if title == "Tarjetas Personales":
            return "Excel histórico + data/tarjetas_9x5/formula_editable_config.json"
        if title == "Tarjetas Postales":
            return "Excel histórico + data/tarjetas_postales/formula_editable_config.json"
        if title == "Folletos":
            return "Excel histórico + data/folletos/formula_editable_config.json"
        return "Copia de DIGITAL sistema 2025.xlsx (lógica histórica/auditoría)"

    def _split_source_cell(self, raw: Any) -> tuple[str, str]:
        if not raw:
            return "desconocida", "desconocida"
        value = str(raw)
        if "!" not in value:
            return "desconocida", value
        sheet, cell = value.split("!", 1)
        return sheet, cell

    def _slug(self, title: str | None) -> str:
        raw = self._clean_text(title or "").lower()
        replacements = {
            " / ": "_",
            " ": "_",
            "á": "a",
            "é": "e",
            "í": "i",
            "ó": "o",
            "ú": "u",
            "ñ": "n",
            "?": "a",
        }
        for source, target in replacements.items():
            raw = raw.replace(source, target)
        return "".join(ch for ch in raw if ch.isalnum() or ch == "_").strip("_")

    def _read_json(self, relative: str) -> dict[str, Any]:
        path = self.project_root / relative
        if not path.exists():
            return {}
        return json.loads(path.read_text(encoding="utf-8-sig"))

    def _git_version_info(self) -> dict[str, Any]:
        version = self._git_command(["git", "describe", "--tags", "--abbrev=0"]) or "sin_tag"
        commit = self._git_command(["git", "rev-parse", "--short", "HEAD"]) or "sin_commit"
        dirty = bool(self._git_command(["git", "status", "--porcelain"]))
        return {
            "version_base_git": version,
            "commit_actual": commit,
            "repo_dirty": dirty,
            "nota_version": "Generado con cambios locales sin commitear" if dirty else "Generado desde versión limpia",
        }

    def _git_command(self, command: list[str]) -> str:
        try:
            result = subprocess.run(
                command,
                cwd=self.project_root,
                check=False,
                capture_output=True,
                text=True,
                timeout=3,
            )
            return result.stdout.strip()
        except Exception:
            return ""

    def _sanitize_workbook_text(self, workbook: Workbook) -> None:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = self._clean_text(cell.value)

    def _clean_text(self, value: str) -> str:
        replacements = {
            "Ã¡": "á",
            "Ã©": "é",
            "Ã­": "í",
            "Ã³": "ó",
            "Ãº": "ú",
            "Ã±": "ñ",
            "Ã": "Á",
            "Ã‰": "É",
            "Ã": "Í",
            "Ã“": "Ó",
            "Ãš": "Ú",
            "Ã‘": "Ñ",
            "Ã¼": "ü",
            "Im?n": "Imán",
            "ImÃ¡n": "Imán",
            "Papel obra/ilustraci?n": "Papel obra/ilustración",
            "f?rmula": "fórmula",
            "f?rmulas": "fórmulas",
            "configuraci?n": "configuración",
            "informaci?n": "información",
            "hist?rico": "histórico",
            "l?gica": "lógica",
            "m?s": "más",
            "m?nimo": "mínimo",
            "m?nimos": "mínimos",
            "seg?n": "según",
            "PegaManÃ­a": "PegaManía",
        }
        cleaned = value
        for source, target in replacements.items():
            cleaned = cleaned.replace(source, target)
        return cleaned

    def _save_report_copy(self, filename: str, payload: bytes) -> None:
        output_dir = self.project_root / "reports" / "exports"
        output_dir.mkdir(parents=True, exist_ok=True)
        (output_dir / filename).write_bytes(payload)

    def _looks_like_header(self, cell) -> bool:
        if cell.row < 2:
            return False
        value = str(cell.value or "").lower()
        common_headers = {"categoria", "producto", "formato", "cantidad_unidades", "precio_total_sin_iva", "precio_sin_iva", "key", "grupo"}
        row_values = {str(ws_cell.value or "").lower() for ws_cell in cell.parent[cell.row]}
        return bool(row_values & common_headers) and value in row_values
